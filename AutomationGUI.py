# -*- coding: utf-8 -*-
#!/usr/bin/env python

# 10/4/2019 Changelog
# Updated icons and colors
# Updated about page

# 10/2/2019 Changelog
# Fixed saveSelectedResultsAction bug
# extended column range to delete as poalrion excel has hidden columns

# 9/28/19 Changelog
# Added feature to save results of selected test cases only
# Added new action to polarionTableContextMenuEvent called saveSelectedResultsAction
# New function called savedSelectedResults to handle action

# 9/16/19 Changelog
# New feature, can add multiple batch run files
# Batch list table implemented on Settings tab to add multiple batche runs

# 8/20/19 Changelog
# Updated Polarion table context menu
# Added Pause with step number feature
# Added option for selecting print output

# 8/14/19 Changelog
# Additional exceptions and traceback information
# Removed dSpace excel and call function edit boxes
# Updated Sequence builder

# 8/1/19 Changelog
# Improved Polarion context menu
# Improved tableview speed
# Added editable comments column
# Added modified column
# Improved tableview add/remove functionality
# Updating polarion steps feature now use polarion excel
# Removed dSPACE excel support
# Added pause step feature

# 7/23/19 Changelog
# Added dark theme
# Fixed total wait time bug
# Added AD blocks conversion
# Improved sw/hw ecu versions support

# 7/11/19 Changelog
# Title and descriptions are shown when test case is clicked
# Fixed bug with saving pickles
# Migrated to Python 3.6

# 7/3/19 Changelog
# Added Polarion authentication check
# Added run test context menu actions
# Profile is now saved as JSON instead of XML

# 6/28/19 Changelog
# Added additional run test context menu actions
# Fixed wait time sorting

# 6/11/19 Changelog
# Parsed Polarion data are now saved on local directory
# Added view for csv log files
# Changed styling to Fusion
# Saving is now threaded
# Reading excel now uses xlrd for speed
# Added wait times column

# 5/28/19 Changelog
# Added loading bar
# Improved signal/slots in threading operations
# Config settings are now saved in registry
# DTC exception hex input are now validated

# 5/13/19 Changelog
# Added toolbar for polarion and dspace tabs
# Improved thread handling between gui and reading excel files
# Added polarion view

# 5/9/19 Changelog
# Added Polarion service support
# Added hyperlink tracking for jiras, etc.

# 5/6/19 Changelog
# Added loading bar
# Added dSpace excel reading
# Added Polarion excel reading
# Added Polarion excel revision number
# Added threading for excel reading


from PyQt5.QtWidgets import QApplication, QMenu, QAction, QMainWindow, QCompleter, QTableWidgetItem, QFileDialog, QMessageBox, QInputDialog, QLineEdit
from PyQt5.QtGui import QPalette, QColor, QIcon, QStandardItemModel, QStandardItem, QCursor, QPixmap
from PyQt5.QtCore import Qt, QTimer, QSettings, QStringListModel, QThread, pyqtSignal
#import xmltodict
from pathlib import Path
from AutomationGUI_ui import *
import os
import sys
#import strings
import csv
import re
import json
import pickle
import xlrd
from datetime import datetime
from openpyxl import load_workbook
from collections import OrderedDict
import traceback
import webbrowser
# from functools import partial # used for calling with args

# polarion related libs
import zeep
import requests
from lxml import etree
import copy
import time

if sys.version_info[0] < 3:
    FileNotFoundError = IOError

debug = False
version = 'v1.6.4'

def debugPrint(msg):
    if debug:
        print(msg)

class App(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(self.__class__, self).__init__()
        self.setupUi(self)

        # create a frameless window without titlebar
        # self.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        self.setWindowIcon(QIcon(':/icon/karmalogo24'))
        # self.screen = QtWidgets.QApplication(sys.argv).primaryScreen()
        # print(self.screen.size())

        # set default xml config file path
        self.configFile = 'C:/DS_Config/config.xml'
        # self.profileFile = 'C:/DS_Config/profile1.xml'

        self.profileFile = ''
        self.configDict = {}
        self.profileDict = {}
        self.runList = []
        self.runDict = {}
        self.excelData = {}
        self.logDict = OrderedDict()
        self.polarionDict = OrderedDict()
        self.polarionWs = None
        self.polarionWb = None
        self.polarionUsername = ''
        self.polarionPassword = ''
        self.variablePoolDict = {}
        self.updateVariablePool = False  # update variable pool in AutomationDesk
        self.defaultProfile = True  # false when a profile is loaded

        # threads
        self.polarionReadExcelThread = None

        # progress bar
        self.progressBarTimer = QTimer()
        self.progressBarTimer.timeout.connect(self.progressBarAnimation)

        # models
        self.versionTableModel = QStandardItemModel()
        self.versionTableView.setModel(self.versionTableModel)
        self.versionTableModel.itemChanged.connect(self.versionTableViewItemChanged)
        self.versionTableHeader = [
            'ECU',
            'Software',
            'Hardware'
        ]
        self.versionTableModel.setHorizontalHeaderLabels(self.versionTableHeader)

        self.batchTableModel = QStandardItemModel()
        self.batchTableView.setModel(self.batchTableModel)
        self.batchTableHeader = [
            'Polarion Excel File',
            'CSV Report Folder'
        ]
        self.batchTableModel.setHorizontalHeaderLabels(self.batchTableHeader)
        self.batchTableView.resizeColumnsToContents()
        self.batchComboBox.setModel(self.batchTableModel)

        self.addSignalModel = QStandardItemModel()
        self.addSignalListView.setModel(self.addSignalModel)

        self.dtcTableModel = QStandardItemModel()
        self.dtcTableView.setModel(self.dtcTableModel)
        self.dtcTableModel.itemChanged.connect(self.dtcTableModelItemChanged)

        self.dtcTableHeader = [
            'Hex Code',
            'Description',
            'Module',
            'OBD',
        ]
        self.dtcTableModel.setHorizontalHeaderLabels(self.dtcTableHeader)

        self.loadSettings()
        self.unsavedChanges = False

        # general tab
        # self.browsePolarionExcelBtn.clicked.connect(self.browsePolarionExcel)
        # self.browseTestCaseExcelBtn.clicked.connect(self.browseTestCaseExcel)
        # self.browseCallFunctionBtn.clicked.connect(self.browseCallFunction)
        # self.browseCsvReportBtn.clicked.connect(self.browseCsvReport)
        self.browseVariablePoolBtn.clicked.connect(self.browseVariablePool)
        # self.testCaseExcelEdit.textChanged.connect(self.useTestCaseFolderForLogs)

        # connections for changes
        # self.polarionExcelEdit.textChanged.connect(self.setUnsavedChanges)
        # self.polarionExcelEdit.textChanged.connect(self.clearPolarionDict)
        # self.testCaseExcelEdit.textChanged.connect(self.SetUnsavedChanges)
        # self.callFunctionEdit.textChanged.connect(self.SetUnsavedChanges)
        # self.csvReportEdit.textChanged.connect(self.setUnsavedChanges)
        self.versionCheckBox.clicked.connect(self.setUnsavedChanges)
        self.variablePoolEdit.textChanged.connect(self.setUnsavedChanges)
        self.dtcExCheckBox.clicked.connect(self.setUnsavedChanges)
        self.callFunctionDebugCheckbox.clicked.connect(self.setUnsavedChanges)
        self.titleCheckBox.clicked.connect(self.setUnsavedChanges)
        self.descriptionCheckBox.clicked.connect(self.setUnsavedChanges)

        self.logRadioBtn0.clicked.connect(self.setUnsavedChanges)
        self.logRadioBtn1.clicked.connect(self.setUnsavedChanges)
        self.logRadioBtn2.clicked.connect(self.setUnsavedChanges)

        self.versionAddButton.clicked.connect(self.versionAddRow)
        self.versionRemoveButton.clicked.connect(self.versionRemoveRow)

        self.batchAddButton.clicked.connect(self.batchAddRow)
        self.batchRemoveButton.clicked.connect(self.batchRemoveRow)
        self.batchClearButton.clicked.connect(self.batchClearAll)

        # polarion excel toolbutton
        self.polarionExcelMenu = QMenu()

        openPolarionExcelFileAction = QAction(QIcon(':/icon/excel'), 'Open Test Case File', self)
        openPolarionExcelFileAction.triggered.connect(lambda: self.openFileInPath(self.getCurrentExcelPath()))
        self.polarionExcelMenu.addAction(openPolarionExcelFileAction)


        openPolarionExcelFolderAction = QAction(QIcon(':/icon/folderOpen'), 'Open Test Case Folder', self)
        openPolarionExcelFolderAction.triggered.connect(lambda: self.openFolderInPath(self.getCurrentExcelPath()))
        self.polarionExcelMenu.addAction(openPolarionExcelFolderAction)

        # self.polarionExcelToolButton.setMenu(self.polarionExcelMenu)

        # test case excel toolbutton
        # testCaseExcelMenu = QMenu()
        #
        # openTestCaseExcelFileAction = QAction(QIcon(':/icon/excel'), 'Open Test Case File', self)
        # openTestCaseExcelFileAction.triggered.connect(lambda: self.openFileInPath(self.testCaseExcelEdit.text()))
        # testCaseExcelMenu.addAction(openTestCaseExcelFileAction)
        #
        # openTestCaseExcelFolderAction = QAction(QIcon(':/icon/folderOpen'), 'Open Test Case Folder', self)
        # openTestCaseExcelFolderAction.triggered.connect(self.openTestCaseExcelFolder)
        # testCaseExcelMenu.addAction(openTestCaseExcelFolderAction)
        #
        # self.testCaseExcelToolButton.setMenu(testCaseExcelMenu)

        # csv report toolbutton
        csvReportFolderMenu = QMenu()

        openCsvReportFolderAction = QAction(QIcon(':/icon/folderOpen'), 'Open Report Folder', self)
        openCsvReportFolderAction.triggered.connect(lambda: self.openFolderInPath(self.csvReportEdit.text()))
        csvReportFolderMenu.addAction(openCsvReportFolderAction)

        useTestCasePathAction = QAction(QIcon(':/icon/folderOpen'), 'Use {TestCaseFolder}\\Logs', self)
        useTestCasePathAction.triggered.connect(self.useTestCaseFolderForLogs)
        csvReportFolderMenu.addAction(useTestCasePathAction)

        # self.csvReportFolderToolButton.setMenu(csvReportFolderMenu)

        # call function toolbutton
        # callFunctionFolderMenu = QMenu()
        #
        # openCallFunctionFolderAction = QAction(QIcon(':/icon/folderOpen'), 'Open Call Function Folder', self)
        # openCallFunctionFolderAction.triggered.connect(lambda: self.openFolderInPath(self.callFunctionEdit.text()))
        # callFunctionFolderMenu.addAction(openCallFunctionFolderAction)
        #
        # self.callFunctionFolderToolButton.setMenu(callFunctionFolderMenu)

        # variable pool toolbutton
        variablePoolMenu = QMenu()

        reloadVariablePoolFileAction = QAction(QIcon(':/icons/reload'), 'Reload Variable Pool', self)
        reloadVariablePoolFileAction.triggered.connect(self.loadVariablePool)
        variablePoolMenu.addAction(reloadVariablePoolFileAction)

        openVariablePoolFileAction = QAction(QIcon(':/icons/open-in-new'), 'Open Variable Pool File', self)
        openVariablePoolFileAction.triggered.connect(lambda: self.openFileInPath(self.variablePoolEdit.text()))
        variablePoolMenu.addAction(openVariablePoolFileAction)

        openVariablePoolFolderAction = QAction(QIcon(':/icons/folder-outline'), 'Open Variable Pool Folder', self)
        openVariablePoolFolderAction.triggered.connect(lambda: self.openFolderInPath(self.variablePoolEdit.text()))
        variablePoolMenu.addAction(openVariablePoolFolderAction)

        self.variablePoolToolButton.setMenu(variablePoolMenu)

        # logging tab
        self.addSignalBtn.clicked.connect(self.addSignal)
        self.addSignalEdit.returnPressed.connect(self.addSignal)
        self.removeSignalBtn.clicked.connect(self.removeSignal)

        self.logRadioBtn0.clicked.connect(self.hideAddSignal)
        self.logRadioBtn1.clicked.connect(self.hideAddSignal)
        self.logRadioBtn2.clicked.connect(self.showAddSignal)
        self.dtcExCheckBox.clicked.connect(self.dtcExToggle)

        # dtc tab

        self.addDtcExBtn.clicked.connect(self.addDtcEx2)

        self.addDtcExEdit.setInputMask('>NN NN NN NN;_')
        self.addDtcExEdit.setMaxLength(11)
        self.addDtcExEdit.returnPressed.connect(self.addDtcEx2)
        self.removeDtcExBtn.clicked.connect(self.removeDtcEx2)

        # polarion tab
        self.polarionReadExcelButton.clicked.connect(self.readPolarionExcel)
        self.polarionTableViewModel = QStandardItemModel()
        self.polarionTableView.setModel(self.polarionTableViewModel)
        self.polarionTableView.clicked.connect(self.polarionTableViewClicked)
        self.polarionTableHeader = ['TestCase', 'Steps', 'Total Wait Time', 'Run Test',
                                    'TestCase Verdict', 'Modified', 'Comments',
                                    'Hyperlinks']
        self.polarionTableViewModel.setHorizontalHeaderLabels(self.polarionTableHeader)


        self.polarionUpdatePassedButton.clicked.connect(lambda: self.updatePolarionVerdicts('Passed'))
        self.polarionUpdateAllButton.clicked.connect(lambda: self.updatePolarionVerdicts('All'))
        self.polarionSaveExcelButton.clicked.connect(self.savePolarionExcel)
        # self.polarionUpdateRevisionButton.clicked.connect(self.udpatePolarionRevision)
        self.updateHyperlinksButton.clicked.connect(self.updateHyperlinks)
        self.updateStepsButton.clicked.connect(self.updatePolarionSteps)
        # self.polarionCopyRunListButton.clicked.connect(self.polarionCopyRunList)

        self.polarionTableView.customContextMenuRequested.connect(self.polarionTableContextMenuEvent)
        self.polarionTableSelectionModel = self.polarionTableView.selectionModel()
        self.polarionTableSelectionModel.selectionChanged.connect(self.polarionTableSelectionChanged)

        self.loadPolarionJson()
        self.loadLogResults()

        self.batchComboBox.currentTextChanged.connect(self.clearPolarionTab)

        # settings tab
        self.updateVariablePoolCheckBox.clicked.connect(self.toggleUpdateVariablePool)

        # Convert sequence tab
        self.sequenceBrowseButton.clicked.connect(self.browseSequenceFile)
        self.sequenceConvertButton.clicked.connect(self.convertSequence)
        # self.tbxCopyTitleButton.clicked.connect(lambda: self.sequenceCopy(self.tbxTitleLineEdit))
        # self.tbxCopyDescButton.clicked.connect(lambda: self.sequenceCopy(self.tbxDescLineEdit))
        # self.tbxCopyTextButton.clicked.connect(lambda: self.sequenceCopy(self.tbxPlainTextEdit))

        # file menu
        self.actionLoad.triggered.connect(self.browseProfile)
        self.actionNew.triggered.connect(self.newProfile)
        self.actionSave.triggered.connect(self.saveProfile)
        self.actionSaveAs.triggered.connect(self.saveAsProfile)
        # self.actionAbout.triggered.connect(lambda: self.tabWidget.setCurrentIndex(5))
        self.actionExit.triggered.connect(self.exit)
        self.actionOpenProfileFolder.triggered.connect(lambda: self.openFolderInPath(str(self.profileFile)))

        # gui layout related
        # dropShadow = QGraphicsDropShadowEffect()
        # dropShadow.setXOffset(1)
        # dropShadow.setYOffset(1)
        # dropShadow.setBlurRadius(6)
        # self.tabWidget.setGraphicsEffect(dropShadow)

        self.tabWidget.setCurrentIndex(0)
        self.logRadioBtn0.setChecked(True)
        self.hideAddSignal()

        self.autorunCheckBox.clicked.connect(self.autoRunClicked)

        self.startTimer = QTimer()
        self.startTimer.timeout.connect(self.tick)

        if self.autorunCheckBox.isChecked():
            self.startTimer.start(1000)
            self.startTimerCount = self.autorunSpinBox.value()

    # def mousePressEvent(self, event):
    #     print(event.button())

    # def resizeEvent(self, event):
    #     self.tabWidget.setFixedHeight(self.height()-160)
    #     self.tabWidget.setFixedWidth(self.width()-40)
    #     self.gridLayout.width
    #     # QtGui.QMainWindow.resizeEvent(self, event)

    def progressBarAnimation(self):
        """Animate progress bar"""
        try:
            value = self.progressBar.value()
            self.progressBar.setValue(0 if value > 99 else value + 20)
        except:
            print(traceback.format_exc())

    def showLoadingBar(self):
        """Show progress bar animation"""
        try:
            self.progressBarTimer.start(1000)
            self.polarionToolbarLayout.setEnabled(False)
            self.setCursor(QCursor(Qt.WaitCursor))
        except:
            print(traceback.format_exc())

    def hideLoadingBar(self):
        """Stop progress bar animation"""
        try:
            self.progressBarTimer.stop()
            self.progressBar.setValue(0)
            self.polarionToolbarLayout.setEnabled(True)
            self.unsetCursor()
        except:
            print(traceback.format_exc())

    def getExcelData(self):
        """Get excel data for AutomationDesk"""
        try:
            polarionDict = self.polarionDict
            self.excelData = {}
            excelData = self.excelData

            for t in polarionDict:
                stepsCompact = []
                steps = polarionDict[t]['steps']

                for i, s in enumerate(steps):
                    row = [str(i + 1)]
                    row.extend(s[5:12])
                    stepsCompact.append(row)

                excelData['{}_Steps'.format(t)] = stepsCompact
        except:
            print(traceback.format_exc())

    def savePolarionDict(self):
        """Save Polarion results to a JSON file"""
        try:
            if len(self.getCurrentExcelPath()) > 0 and len(self.polarionDict) > 0:

                polarionJsonPath = self.getCurrentExcelPath().replace('.xlsx','.json')

                with open(polarionJsonPath, 'w') as f:
                    polarionJson = json.dumps(self.polarionDict, sort_keys=True, indent=4)
                    f.write(polarionJson)
        except:
            print(traceback.format_exc())

    def loadPolarionJson(self):
        """Load polarion results from a JSON file"""

        polarionJsonPath = self.getCurrentExcelPath().replace('.xlsx', '.json')
        # if there is an existing json file for the polarion excel, load up data from json
        if os.path.exists(polarionJsonPath):
            with open(polarionJsonPath) as f:
                self.polarionDict = json.load(f)
                self.updatePolarionTableModel()

    def getCurrentExcelPath(self):
        return self.batchComboBox.currentText()

    def getCurrentCsvReportFolder(self):
        header = self.batchTableHeader
        model = self.batchTableModel
        combo = self.batchComboBox
        csvReportFolderCol = header.index('CSV Report Folder')
        return model.item(combo.currentIndex(), csvReportFolderCol).text()

    def saveLogResults(self):
        """Save log file results to a pickle file"""
        try:
            # print(csvReportFolder)
            logPicklePath = os.path.join(self.getCurrentCsvReportFolder(), 'LogResults.pkl')
            with open(logPicklePath, 'wb') as f:
                pickle.dump(self.logDict, f)
        except:
            print(traceback.format_exc())

    def loadLogResults(self):
        """Load log file results if it exists"""
        try:
            logPicklePath = os.path.join(self.getCurrentCsvReportFolder(), 'LogResults.pkl')
            try:
                with open(logPicklePath, 'rb') as f:
                    self.logDict = pickle.load(f)
            except:
                print('Exception for loadLogResults')
        except:
            print(traceback.format_exc())

    def markAllModified(self, all):
        """Set modified column for all"""
        try:
            model = self.polarionTableViewModel
            header = self.polarionTableHeader
            modifiedCol = header.index('Modified')
            rowCount = model.rowCount()

            for i in range(0, rowCount):
                modifiedItem = model.item(i, modifiedCol)
                modifiedItem.setCheckState(Qt.Checked if all else Qt.Unchecked)
        except:
            print(traceback.format_exc())

    def markSelectedModified(self, modified):
        """Mark selected testcases as modified"""
        try:
            model = self.polarionTableViewModel
            view = self.polarionTableView
            header = self.polarionTableHeader
            selectedIndexes = view.selectedIndexes()

            if self.polarionTableView.model():
                modifiedCol = header.index('Modified')

                selectedItems = [model.item(x.row(), modifiedCol) for x in selectedIndexes]

                for modifiedItem in selectedItems:
                    modifiedItem.setCheckState(Qt.Checked if modified else Qt.Unchecked)
        except:
            print(traceback.format_exc())

    def runSelected(self, run):
        """Run selected test cases"""
        try:
            model = self.polarionTableViewModel
            view = self.polarionTableView
            header = self.polarionTableHeader
            selectedIndexes = view.selectedIndexes()

            if len(selectedIndexes) > 0:
                runTestCol = header.index('Run Test')

                selectedItems = [model.item(x.row(), runTestCol) for x in selectedIndexes]

                for runTestItem in selectedItems:
                    runTestItem.setCheckState(Qt.Checked if run else Qt.Unchecked)

        except:
            print(traceback.format_exc())

    def runAll(self, all):
        """Run all test cases"""
        try:
            model = self.polarionTableViewModel
            header = self.polarionTableHeader
            runTestCol = header.index('Run Test')

            for row in range(0, model.rowCount()):
                runTestItem = model.item(row, runTestCol)
                runTestItem.setCheckState(Qt.Checked if all else Qt.Unchecked)
        except:
            print(traceback.format_exc())

    def saveSelectedResults(self):
        """Saves the result of selected test cases only"""
        class saveSelectedResultsThread(QThread):
            showLoadingBarSignal = pyqtSignal()
            hideLoadingBarSignal = pyqtSignal()
            appendMessageSignal = pyqtSignal('PyQt_PyObject')
            updatePolationRevisionSignal = pyqtSignal()

            def __init__(self):
                QThread.__init__(self)

            def run(self):
                try:
                    if len(self.filePath) > 0:
                        self.showLoadingBarSignal.emit()
                        try:
                            # create new polarion excel file with results of only selected test cases

                            # close any existing polarion excel file
                            if self.polarionWb:
                                self.polarionWb.close()

                            # open desired polarion file
                            self.polarionWb = load_workbook(filename=self.polarionExcelPath)
                            self.polarionWs = self.polarionWb.get_sheet_by_name('Sheet1')

                            # delete everything after the first row on Sheet 1
                            max_row = self.polarionWs.max_row
                            max_col = self.polarionWs.max_column + 1
                            for i in range(2,max_row):
                                for j in range(1,max_col):
                                    del self.polarionWs._cells[i,j]

                            # write into the excel sheet only the selected test cases
                            self.polarionWs._current_row = 1
                            for t in self.selectedTestCases:
                                if self.polarionDict[t]['length'] == self.logDict[t]['length']:
                                    if self.polarionDict[t]['testCaseVerdict'] in ['Passed', 'Deferred', 'Error']:
                                        stepsLength = self.polarionDict[t]['length']
                                        # iterate through each step of a test case and add row to the excel file
                                        for i in range(stepsLength):
                                            # use polarionDict for col 0 - 10, 16, 17 on the excel file
                                            row_content = self.polarionDict[t]['steps'][i]

                                            # use logDict for filling out col for actual result, step verdict,
                                            row_content[12] = self.logDict[t]['actualResult'][i]    # actual results col
                                            row_content[13] = self.logDict[t]['passList'][i]        # step verdict col

                                            if i == 0:
                                                row_content[14] = self.logDict[t]['testCaseVerdict'] # test case verdict col
                                                row_content[15] = self.logDict[t]['testComment']# test comment col

                                            # add the step to the polarion excel file
                                            self.polarionWs.append(row_content)

                            self.polarionWb.save(self.filePath)
                            self.appendMessageSignal.emit('Save successful')

                        except KeyError:
                            pass
                        except IOError as error:
                            if str(error).__contains__('Permission denied'):
                                self.appendMessageSignal.emit('Permission Denied. Unable to save excel file')
                except:
                    print(traceback.format_exc())
                finally:
                    self.hideLoadingBarSignal.emit()


        try:
            model = self.polarionTableViewModel
            view = self.polarionTableView
            header = self.polarionTableHeader
            testCaseCol = header.index('TestCase')
            selectedIndexes = view.selectedIndexes()

            # create a list of test case ID numbers based on the selected test cases
            selectedTestCases = []
            if len(selectedIndexes) > 0:
                for eachIndex in selectedIndexes:
                    testCaseItem = model.item(eachIndex.row(), testCaseCol)
                    testCase = testCaseItem.text()
                    selectedTestCases.append(testCase)

            folder = os.path.dirname(self.getCurrentExcelPath())
            filePath, fileType = QFileDialog.getSaveFileName(
                self,
                "Save Polarion Excel File",
                folder,
                'XLSX Files (*.xlsx);;All Files (*)'
            )

            self.saveSelectedResultsThreadObject = saveSelectedResultsThread()
            myThread = self.saveSelectedResultsThreadObject
            myThread.filePath = filePath
            myThread.polarionExcelPath = self.getCurrentExcelPath()
            myThread.polarionWb = self.polarionWb
            myThread.polarionWs = self.polarionWs
            myThread.polarionDict = self.polarionDict
            myThread.logDict = self.logDict
            myThread.selectedTestCases = selectedTestCases
            myThread.showLoadingBarSignal.connect(self.showLoadingBar)
            myThread.hideLoadingBarSignal.connect(self.hideLoadingBar)
            myThread.appendMessageSignal.connect(self.appendPolarionLog)
            #myThread.updatePolarionRevisionSignal.connect(self.udpatePolarionRevision)
            myThread.start()
        except:
            print(traceback.format_exc())

    def makeRunList(self):
        """Scan the run test checkboxes and make the run list/dict"""
        try:
            header = self.polarionTableHeader
            testCaseCol = header.index('TestCase')
            runTestCol = header.index('Run Test')
            model = self.polarionTableViewModel
            rowCount = model.rowCount()

            self.runList = []
            runList = self.runList
            runDict = self.runDict

            for i in range(0, rowCount):
                testCaseItem = model.item(i, testCaseCol)
                runTestItem = model.item(i, runTestCol)
                testCase = testCaseItem.text()
                runTest = runTestItem.text()
                runList.append([testCase, runTest])

                try:
                    self.polarionDict[testCase]['run'] = runTest
                except KeyError:
                    pass

                try:
                    runDict[testCase] = runTest
                except:
                    pass

            def sortfunc(x):
                return x[0]

            runList.sort(key=sortfunc)
        except:
            print(traceback.format_exc())

    def polarionTableContextMenuEvent(self, qpoint):
        """Show context menu for Polarion tableview"""
        try:
            menu = QMenu(self)
            view = self.polarionTableView
            selectedIndexes = view.selectedIndexes()

            if len(selectedIndexes) > 0:
                saveSelectedResultsAction = QAction(QIcon(':/icons/content-save-outline'),'Save Selected Results', self)
                saveSelectedResultsAction.triggered.connect(self.saveSelectedResults)
                menu.addAction(saveSelectedResultsAction)

                copyAction = QAction(QIcon(':/icons/content-copy'), 'Copy TestCase ID(s)', self)
                copyAction.triggered.connect(self.copyTestCaseId)
                menu.addAction(copyAction)

                menu.addSeparator()

                openLinkAction = QAction(QIcon(':/icons/link'), 'Open Hyperlink(s)', self)
                openLinkAction.triggered.connect(self.openHyperlink)
                menu.addAction(openLinkAction)

                openLogFileAction = QAction(QIcon(':/icons/open-in-new'), 'Open Log File(s)', self)
                openLogFileAction.triggered.connect(self.openLogFile)
                menu.addAction(openLogFileAction)

                menu.addSeparator()

                # checkAllAction = QAction(QIcon(':icon_white/checked'), 'Check All', self)
                # checkAllAction.triggered.connect(self.checkAllPolarionTable)
                # menu.addAction(checkAllAction)
                #
                # decheckAllAction = QAction(QIcon(':icon_white/notChecked'), 'Check None', self)
                # decheckAllAction.triggered.connect(self.decheckAllPolarionTable)
                # menu.addAction(decheckAllAction)

                modifySelectedAction = QAction(QIcon(':/icons/checkbox-marked-outline'), 'Mark as modified', self)
                modifySelectedAction.triggered.connect(lambda: self.markSelectedModified(True))
                menu.addAction(modifySelectedAction)

                modifyAllAction = QAction(QIcon(':/icons/check-box-multiple-outline'), 'Mark All as modified', self)
                modifyAllAction.triggered.connect(lambda: self.markAllModified(True))
                menu.addAction(modifyAllAction)

                menu.addSeparator()

                unmodifySelectedAction = QAction(QIcon(':/icons/checkbox-blank-outline'), 'Mark as unmodified', self)
                unmodifySelectedAction.triggered.connect(lambda: self.markSelectedModified(False))
                menu.addAction(unmodifySelectedAction)

                unmodifyAllAction = QAction(QIcon(':/icons/checkbox-multiple-blank-outline'), 'Mark All as unmodified', self)
                unmodifyAllAction.triggered.connect(lambda: self.markAllModified(False))
                menu.addAction(unmodifyAllAction)

                menu.addSeparator()

                runSelectedTestCasesAction = QAction(QIcon(':/icons/checkbox-marked-outline'), 'Run selected', self)
                runSelectedTestCasesAction.triggered.connect(lambda: self.runSelected(True))
                menu.addAction(runSelectedTestCasesAction)

                runAllTestCasesAction = QAction(QIcon(':/icons/check-box-multiple-outline'), 'Run All', self)
                runAllTestCasesAction.triggered.connect(lambda: self.runAll(all=True))
                menu.addAction(runAllTestCasesAction)

                menu.addSeparator()

                dontRunSelectedTestCasesAction = QAction(QIcon(':/icons/checkbox-blank-outline'), 'Do Not run selected', self)
                dontRunSelectedTestCasesAction.triggered.connect(lambda: self.runSelected(False))
                menu.addAction(dontRunSelectedTestCasesAction)

                dontRunAllTestCasesAction = QAction(QIcon(':/icons/checkbox-multiple-blank-outline'), 'Run None', self)
                dontRunAllTestCasesAction.triggered.connect(lambda: self.runAll(all=False))
                menu.addAction(dontRunAllTestCasesAction)

                menu.addSeparator()


                if menu.actions().__len__() > 0:
                    menu.popup(QCursor.pos())
        except:
            print(traceback.format_exc())

    def polarionTableViewClicked(self, qmodelindex):
        """Handler for Polarion tableview clicked event"""
        try:
            header = self.polarionTableHeader
            row = qmodelindex.row()
            column = qmodelindex.column()

            testCaseCol = header.index('TestCase')
            runTestCol = header.index('Run Test')
            testCaseId = self.polarionTableViewModel.item(row, testCaseCol).text()

            try:
                csvFilePath = self.logDict[testCaseId]['filePath']
                if os.path.exists(csvFilePath):
                    self.updateLogTableWidget(csvFilePath)
                    self.logTableWidget.scrollToTop()
            except:
                self.logTableWidget.clear()
                pass

            try:
                title = self.polarionDict[testCaseId]['title']
                description = self.polarionDict[testCaseId]['desc']
                self.titleEdit.setText(title)
                self.descriptionEdit.setText(description)
            except KeyError:
                print('KeyError for def polarionTableViewClicked, title & description')
        except:
            print(traceback.format_exc())

    def polarionTableViewModelItemChanged(self, item):
        """Handler for Polarion model item changed event"""
        try:
            row = item.row()
            column = item.column()
            model = self.polarionTableViewModel
            polarionDict = self.polarionDict
            testcaseCol = self.polarionTableHeader.index('TestCase')
            runTestCol = self.polarionTableHeader.index('Run Test')
            commentsCol = self.polarionTableHeader.index('Comments')
            modifiedCol = self.polarionTableHeader.index('Modified')

            testcaseItem = model.item(row, testcaseCol)
            testcase = testcaseItem.text()
            run = False

            if column == runTestCol:
                run = item.checkState() == Qt.Checked
                item.setText(str(run))
                try:
                    polarionDict[testcase]['run'] = run
                except:
                    pass

            if column == commentsCol:
                comments = item.text()
                if 'comments' not in polarionDict[testcase]:
                    polarionDict[testcase]['comments'] = ''
                try:
                    polarionDict[testcase]['comments'] = comments
                except:
                    pass

            if column == modifiedCol:
                checkedState = item.checkState() == Qt.Checked
                item.setText(str(checkedState))
                if 'modified' not in polarionDict[testcase]:
                    polarionDict[testcase]['modified'] = False
                try:
                    polarionDict[testcase]['modified'] = checkedState
                except:
                    pass
        except:
            print(traceback.format_exc())

    def polarionTableSelectionChanged(self):
        """Show number of testcases that are selected"""
        try:
            view = self.polarionTableView
            if view.model():
                self.selectedEdit.setText(view.selectedIndexes().__len__().__str__())
        except:
            print(traceback.format_exc())

    def clearPolarionTab(self, item):
        self.polarionDict.clear()
        self.polarionTableViewModel.clear()
        self.titleEdit.clear()
        self.descriptionEdit.clear()
        self.logTableWidget.clear()
        self.selectedEdit.clear()
        self.polarionRevisionLineEdit.clear()
        self.polarionOverViewLineEdit.clear()
        self.polarionLogEdit.clear()

    def updateLogTableWidget(self, csvFilePath):
        """Update the log table view when clicking on test case"""
        try:
            self.logTableHeader = ['Action', 'Description', 'Variable', 'Settings', 'Value', 'Wait', 'Actual Value', 'Step Verdict', 'Step Message']
            header = self.logTableHeader

            # stepCol = header.index('Step')
            actionCol = header.index('Action')
            descriptionCol = header.index('Description')
            variableCol = header.index('Variable')
            settingsCol = header.index('Settings')
            valueCol = header.index('Value')
            waitCol = header.index('Wait')
            actualCol = header.index('Actual Value')
            verdictCol = header.index('Step Verdict')
            statusCol = header.index('Step Message')

            self.logTableWidget.clear()

            with open(csvFilePath) as csvFile:
                csvReader = csv.DictReader(csvFile)
                allRows = [row for row in csvReader]

                self.logTableWidget.setRowCount(allRows.__len__())
                self.logTableWidget.setColumnCount(header.__len__())

                rowIndex = 0

                for row in allRows:
                    try:
                        self.logTableWidget.setItem(rowIndex, actionCol, QTableWidgetItem(row['Action']))
                    except KeyError:
                        pass

                    try:
                        self.logTableWidget.setItem(rowIndex, descriptionCol, QTableWidgetItem(row['Description']))
                    except KeyError:
                        pass

                    try:
                        self.logTableWidget.setItem(rowIndex, variableCol, QTableWidgetItem(row['Variable']))
                    except KeyError:
                        pass

                    try:
                        self.logTableWidget.setItem(rowIndex, settingsCol, QTableWidgetItem(row['Settings']))
                    except KeyError:
                        pass

                    try:
                        self.logTableWidget.setItem(rowIndex, valueCol, QTableWidgetItem(row['Value']))
                    except KeyError:
                        pass

                    try:
                        self.logTableWidget.setItem(rowIndex, waitCol, QTableWidgetItem(row['Wait']))
                    except KeyError:
                        pass

                    try:
                        self.logTableWidget.setItem(rowIndex, actualCol, QTableWidgetItem(row['Actual Value']))
                    except KeyError:
                        pass

                    try:
                        verdict = row['Step Verdict']
                        verdictItem = QTableWidgetItem(verdict)
                        if verdict == 'Passed':
                            verdictItem.setIcon(QIcon(':/icon/passed'))
                        elif verdict == 'Deferred' or verdict == 'Error' or verdict == 'Not Passed':
                            verdictItem.setIcon(QIcon(':/icon/failed'))

                        self.logTableWidget.setItem(rowIndex, verdictCol, verdictItem)

                    except KeyError:
                        print('KeyError for updateLogTableWidget, verdict')

                    try:
                        self.logTableWidget.setItem(rowIndex, statusCol, QTableWidgetItem(row['Step Message']))
                    except KeyError:
                        print('KeyError for updateLogTableWidget, QTableWidgetItem')

                    rowIndex += 1

                self.logTableWidget.setHorizontalHeaderLabels(header)
                self.logTableWidget.setAlternatingRowColors(True)
                self.logTableWidget.resizeColumnsToContents()
                self.logTableWidget.setColumnWidth(descriptionCol, 20)
        except:
            print(traceback.format_exc())

    def openLogFile(self):
        """Open log file in system's default editor"""
        try:
            view = self.polarionTableView
            header = self.polarionTableHeader
            model = self.polarionTableViewModel
            selectedIndexes = view.selectedIndexes()
            testCaseIdCol = header.index('TestCase')

            if len(selectedIndexes)>0:
                openList = []
                selectedItems = [model.item(x.row(), testCaseIdCol) for x in selectedIndexes]

                for item in selectedItems:
                    testcase = item.text()
                    filePath = self.logDict[testcase]['filePath']
                    if os.path.exists(filePath):
                        os.startfile(filePath)
        except:
            print(traceback.format_exc())

    def copyTestCaseId(self):
        """Copy selection in Polarion table"""
        try:
            model = self.polarionTableViewModel
            view = self.polarionTableView
            selectedIndexes = view.selectedIndexes()
            testcaseCol = self.polarionTableHeader.index('TestCase')
            selectedItems = [model.item(x.row(), testcaseCol) for x in selectedIndexes]
            copyList = []

            for item in selectedItems:
                testcase = item.text()
                copyList.append(testcase)

            copyString = '\n'.join(copyList)

            cb = QApplication.clipboard()
            cb.clear(mode=cb.Clipboard)
            cb.setText(copyString, mode=cb.Clipboard)
        except:
            print(traceback.format_exc())

    def openHyperlink(self):
        """Open hyperlink in system's default web browser"""
        try:
            model = self.polarionTableViewModel
            view = self.polarionTableView
            selectedIndexes = view.selectedIndexes()
            hyperlinkCol = self.polarionTableHeader.index('Hyperlinks')
            selectedItems = [model.item(x.row(), hyperlinkCol) for x in selectedIndexes]

            for item in selectedItems:
                link = item.text()
                if len(link) > 0:
                    linkSplit = link.split(',')
                    for l in linkSplit:
                        webbrowser.open_new_tab(l)
        except:
            print(traceback.format_exc())

    def autoRunClicked(self):
        """Handler for autorun checkbox clicked event"""
        try:
            if self.autorunCheckBox.isChecked():
                self.startTimer.start(1000)
                self.startTimerCount = self.autorunSpinBox.value()
            else:
                self.startTimer.stop()
                self.autorunCheckBox.setText('Autorun')
        except:
            print(traceback.format_exc())

    def tick(self):
        """Ticker for autorun timer"""
        try:
            if self.startTimerCount > 0:
                self.autorunCheckBox.setText('Autorun in {}s'.format(self.startTimerCount))
                # self.exitButton.setText('Continue ({})'.format(self.startTimerCount))
                debugPrint(self.startTimerCount)
                self.startTimerCount = self.startTimerCount - 1
            else:
                self.startTimer.stop()
                debugPrint('Timer stop')
                self.autorunCheckBox.setText('Autorun')
                # self.exitButton.setText('Continue')
                if self.autorunCheckBox.isChecked():
                    self.exit()
        except:
            print(traceback.format_exc())

    def toggleUpdateVariablePool(self):
        """Toggle update variablepool checkbox"""
        try:
            self.updateVariablePool = self.updateVariablePoolCheckBox.isChecked()
            debugPrint('UpdateVariablePool={}'.format(self.updateVariablePool))
        except:
            print(traceback.format_exc())

    def openFolderInPath(self, path):
        """Open folder contained in path"""
        try:
            if os.path.exists(path):
                if os.path.isdir(path):
                    os.startfile(path)
                else:
                    os.startfile(os.path.dirname(path))
        except:
            print(traceback.format_exc())

    def openFileInPath(self, path):
        """Open file contained in path"""
        try:
            if os.path.exists(path):
                os.startfile(path)
        except:
            print(traceback.format_exc())

    def useTestCaseFolderForLogs(self):
        """Use test case folder for logs"""
        try:
            testCaseExcelPath = Path(self.getCurrentExcelPath())
            dirName = os.path.dirname(str(testCaseExcelPath))
            newCsvReportFolder = os.path.join(dirName, 'Logs')

            if os.path.exists(dirName):
                msgReply = QMessageBox.question(
                    self,
                    'CSV Report Folder',
                    'Use \'{}\' to store reports?'.format(newCsvReportFolder),
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )

                if msgReply == QMessageBox.Yes:
                    self.csvReportEdit.setText(str(newCsvReportFolder))
                    if not os.path.exists(newCsvReportFolder):
                        try:
                            os.mkdir(newCsvReportFolder)
                        except:
                            pass
        except:
            print(traceback.format_exc())

    def hideAddSignal(self):
        """Hide add signal groupbox"""
        try:
            self.addSignalGroupBox.hide()
        except:
            print(traceback.format_exc())

    def showAddSignal(self):
        """Show add signal groupbox"""
        try:
            self.addSignalGroupBox.show()
        except:
            print(traceback.format_exc())

    def dtcExToggle(self):
        """Toggle DTC exception checkbox"""
        try:
            self.showDtcException(self.dtcExCheckBox.isChecked())
        except:
            print(traceback.format_exc())

    def showDtcException(self, visible=True):
        """Show DTC exception groubox"""
        try:
            if visible:
                self.dtcExGroupBox.show()
            else:
                self.dtcExGroupBox.hide()
        except:
            print(traceback.format_exc())

    def setUnsavedChanges(self):
        """Set unsaved changes flag"""
        try:
            self.unsavedChanges = True
            self.statusbar.showMessage('You have unsaved changes')
        except:
            print(traceback.format_exc())

    def newProfile(self):
        """Create new profile and clearing edits boxes, tables, etc."""
        try:
            self.setTitle('New Profile')
            self.profileFile = ''
            # self.testCaseExcelEdit.clear()
            self.polarionExcelEdit.clear()
            self.csvReportEdit.clear()
            # self.callFunctionEdit.clear()
            self.variablePoolEdit.clear()
            self.versionTableModel.clear()
            self.versionCheckBox.setChecked(False)
            self.callFunctionDebugCheckbox.setChecked(False)
            self.dtcExCheckBox.setChecked(False)
            self.dtcTableModel.clear()
            self.clearPolarionTab()
            self.statusbar.showMessage('Default profile loaded')
        except:
            print(traceback.format_exc())

    def browseProfile(self):
        """Browse for a profile"""
        try:
            profileFolder = os.path.dirname(self.profileFile)
            filePath, fileType = QFileDialog.getOpenFileName(
                self,
                "Open Profile",
                profileFolder,
                "JSON Files (*.json);;All Files (*)"
            )

            if filePath:
                self.profileFile = filePath
                self.loadProfile()
                self.saveSettings()
        except:
            print(traceback.format_exc())

    def browsePolarionExcel(self):
        """Browse for Polarion excel file"""
        try:
            filePath = self.browseFile(self.getCurrentExcelPath(), 'Select Polarion TestCase Excel File', 'Excel Files (*.xlsx)')
            self.polarionExcelEdit.setText(filePath)

            # clear when browsing for new file
            self.useTestCaseFolderForLogs()
            self.polarionDict.clear()
            self.clearPolarionTab()

        except:
            print(traceback.format_exc())

    def browseCsvReport(self):
        """Browse for csv report folder"""
        try:
            self.csvReportEdit.setText(self.browseFolder(self.csvReportEdit.text(), 'Select CSV Report Directory'))
        except:
            print(traceback.format_exc())

    def browseVariablePool(self):
        """"Browse for variable pool file and load"""
        try:
            filePath = self.browseFile(self.variablePoolEdit.text(), 'Open Variable Pool', 'TXT Files (*.txt)')
            self.variablePoolEdit.setText(filePath)
            self.loadVariablePool()
        except:
            print(traceback.format_exc())

    def browseFile(self, filePath='', titleDialog='', fileType=''):
        """Browse for file"""
        try:
            if len(filePath) > 0:
                folder = str(os.path.dirname(filePath))
            else:
                folder = ''

            openFilePath, openFileType = QFileDialog.getOpenFileName(
                self,
                titleDialog,
                folder,
                "{};;All Files (*)".format(fileType)
            )
            return openFilePath.replace('/', '\\') if len(openFilePath) > 0 else filePath
        except:
            print(traceback.format_exc())

    def browseFolder(self, folder, title):
        """Browse for a folder"""
        try:
            folderPath = QFileDialog.getExistingDirectory(self, title, folder)
            return folderPath.replace('/', '\\') if len(folderPath) > 0 else folder
        except:
            print(traceback.format_exc())

    def browseSequenceFile(self):
        try:
            filePath = self.browseFile(self.sequenceFileEdit.text(), 'Select Sequence File', 'SQX Files (*.sqx)')
            self.sequenceFileEdit.setText(filePath)
        except:
            print(traceback.format_exc())

    def convertSequence(self):
        """Convert AD sequence to tab delimited text"""
        try:
            import xmltodict

            txtEdit = self.sequenceEdit
            txtEdit.clear()

            with open(self.sequenceFileEdit.text(), 'r') as f:
                xml_input = f.read()
                testCase = xmltodict.parse(xml_input)

                sequence = testCase['Standard.Sequence']
                subsystems = sequence['SubSystems']
                block = subsystems['Block']
                subsystems2 = block['SubSystems']
                block2 = subsystems2['Block']
                inheritedDataObjects = [x['InheritedDataObjects']['DOB'] for x in block2]
                steps = []
                stepString = ''

                def checkValue(s):
                    if isinstance(s, str):
                        s = s.strip('\'')
                        return '-' if len(s) == 0 else s
                    return str(s)

                for i, cBlock in enumerate(block2):
                    blockLink = cBlock['@CustomLibraryLink'].strip('\'')
                    blockName = blockLink.split('.')[-1]
                    blockEnable = eval(cBlock['@EnableMode'])

                    if blockName in ['InitMAPort', 'ReleaseMAPort']:
                        continue

                    if blockEnable:
                        dob = cBlock['InheritedDataObjects']['DOB']
                        dataObjectsNames = [x['@Name'].strip('\'') for x in dob]
                        desc = dob[dataObjectsNames.index('Descrip')]['@Value']
                        variable = dob[dataObjectsNames.index('Variable')]['@Value']
                        settings = dob[dataObjectsNames.index('Settings')]['@Value']
                        value = dob[dataObjectsNames.index('Value')]['@Value']
                        wait = dob[dataObjectsNames.index('Wait')]['@Value']
                        remarks = dob[dataObjectsNames.index('Remarks')]['@Value']

                        stepList = [checkValue(x) for x in
                                    [blockName, desc, variable, settings, value, wait, remarks]]
                        # print('\t'.join(stepList))
                        # txtEdit.appendPlainText('\t'.join(stepList))
                        steps.append('\t'.join(stepList))
                        # stepString += '\t'.join(stepList) + '\n'

                stepString = '\n'.join(steps)
                txtEdit.appendPlainText(stepString)
                cb = QApplication.clipboard()
                cb.clear(mode=cb.Clipboard)
                cb.setText(stepString, mode=cb.Clipboard)
                self.statusbar.showMessage('Sequence copied to clipboard')

        except:
            print(traceback.format_exc())

        # try:
        #     with open(self.tbxLineEdit.text(), 'r') as f:
        #         xml_input = f.read()
        #
        #         testCase = xmltodict.parse(xml_input)
        #
        #         dob = testCase['TemplateBlock']['Block']['DataObjects']['DOB']
        #         titleDescNames = [x['@Name'].strip('\'') for x in dob]
        #         title = dob[titleDescNames.index('Title')]['@Value'].strip('\'')
        #         desc = dob[titleDescNames.index('Desc')]['@Value'].strip('\'')
        #
        #         self.tbxTitleLineEdit.setText(title)
        #         self.tbxDescLineEdit.setText(desc)
        #
        #         blocks = testCase['TemplateBlock']['Block']['SubSystems']['Block']
        #
        #         for i, block in enumerate(blocks):
        #             blockLink = block['@CustomLibraryLink'].strip('\'')
        #             blockName = blockLink.split('')[-1]
        #             blockEnable = block['@EnableMode']
        #
        #             dataObjects = block['InheritedDataObjects']['DOB']
        #             dataObjectsNames = [x['@Name'].strip('\'') for x in block['InheritedDataObjects']['DOB']]
        #
        #             action = dataObjects[dataObjectsNames.index('Action')]['@Value']
        #             desc = dataObjects[dataObjectsNames.index('Descrip')]['@Value']
        #             variable = dataObjects[dataObjectsNames.index('Variable')]['@Value']
        #             settings = dataObjects[dataObjectsNames.index('Settings')]['@Value']
        #             value = dataObjects[dataObjectsNames.index('Value')]['@Value']
        #             wait = dataObjects[dataObjectsNames.index('Wait')]['@Value']
        #             remarks = dataObjects[dataObjectsNames.index('Remarks')]['@Value']
        #
        #             def checkValue(s):
        #                 if isinstance(s, str):
        #                     s = s.strip('\'')
        #                     return '-' if len(s) == 0 else s
        #                 return str(s)
        #
        #             if eval(blockEnable):
        #                 stepList = [checkValue(x) for x in [blockName, desc, variable, settings, value, wait, remarks]]
        #                 txtEdit.appendPlainText('\t'.join(stepList))
        # except KeyError as e:
        #     print(e)

    def saveProfile(self):
        """Save current profile as a json"""
        try:
            if self.profileFile == '':
                filePath, fileType = QFileDialog.getSaveFileName(
                    self,
                    "Save Profile As",
                    '',
                    "JSON Files (*.json);;All Files (*)"
                )

                if filePath:
                    self.profileFile = filePath

            # save config file as well
            # self.makeRunList()
            self.saveSettings()
            self.savePolarionDict()

            # updates the profile dict before dumping to json
            self.updateProfileFromGui()

            try:
                with open(self.profileFile, 'w') as f:
                    jsonDump = json.dumps(self.profileDict, sort_keys=True, indent=4)
                    f.write(jsonDump)
                    self.setTitle(self.profileFile)
                    # self.defaultProfile = False
                    self.unsavedChanges = False
                    self.statusbar.showMessage('Saving Profile...OK!')

            except FileNotFoundError:
                self.statusbar.showMessage('Unable to save profile')
            except:
                print(traceback.format_exc())
        except:
            print(traceback.format_exc())

    def saveAsProfile(self):
        """Save as current profile as a JSON file"""
        try:
            profileFolder = os.path.dirname(self.profileFile)
            filePath, fileType = QFileDialog.getSaveFileName(
                self,
                "Save Profile As",
                profileFolder,
                "JSON Files (*.json);;All Files (*)"
            )

            if filePath:
                self.profileFile = filePath
                self.saveProfile()
        except:
            print(traceback.format_exc())

    def updateProfileFromGui(self):
        """Updates the profile dict from GUI elements"""
        try:
            # generate the log mode number based on checkboxes
            logMode = (self.logRadioBtn0.isChecked() + self.logRadioBtn1.isChecked() * 2 + self.logRadioBtn2.isChecked() * 3) - 1

            # update signal list
            signalList = []
            for i in range(0, self.addSignalModel.rowCount()):
                signalList.append(self.addSignalModel.item(i, 0).text())

            dtcExList = []
            dtcModel = self.dtcTableModel
            dtcHexCodeCol = self.dtcTableHeader.index('Hex Code')
            dtcDescCol = self.dtcTableHeader.index('Description')
            dtcModuleCol = self.dtcTableHeader.index('Module')
            dtcOBDCol = self.dtcTableHeader.index('OBD')

            for row in range(0, dtcModel.rowCount()):
                hexCode = dtcModel.item(row, dtcHexCodeCol).text()
                desc = dtcModel.item(row, dtcDescCol).text()
                module = dtcModel.item(row, dtcModuleCol).text()
                obd = dtcModel.item(row, dtcOBDCol).checkState() == Qt.Checked
                dtcExList.append({
                    'Hex Code': hexCode,
                    'Description': desc,
                    'Module': module,
                    'OBD': obd,
                })

            versionDict = {}
            versionModel = self.versionTableModel
            versionHeader = self.versionTableHeader

            swCol = versionHeader.index('Software')
            hwCol = versionHeader.index('Hardware')
            ecuCol = versionHeader.index('ECU')

            for row in range(0, versionModel.rowCount()):
                sw = ''
                hw = ''
                try:
                    sw = versionModel.item(row, swCol).text()
                except:
                    pass
                try:
                    hw = versionModel.item(row, hwCol).text()
                except:
                    pass

                try:
                    ecu = versionModel.item(row, ecuCol).text()
                    versionDict[ecu] = {
                        'Software': sw,
                        'Hardware': hw,
                    }
                except:
                    pass

            # batchDict = {}
            # batchModel = self.batchTableModel
            # batchHeader = self.batchTableHeader
            #
            # polarionExcelCol = batchHeader.index('Polarion Excel File')
            # csvReportCol = batchHeader.index('CSV Report Folder')

            # for row in range(0, batchModel.rowCount()):
            #     csvReport = ''
            #     try:
            #         csvReport = batchModel.item(row, csvReportCol).text()
            #     except:
            #         pass
            #     try:
            #         polarionExcel = batchModel.item(row, polarionExcelCol).text()
            #         batchDict[polarionExcel] = {
            #             'CSV Report Folder': csvReport
            #         }
            #     except:
            #         pass

            batchList = [] # list of dictionaries
            batchModel = self.batchTableModel
            batchHeader = self.batchTableHeader
            polarionExcelCol = batchHeader.index('Polarion Excel File')
            csvReportFolderCol = batchHeader.index('CSV Report Folder')

            for row in range(0, batchModel.rowCount()):
                polarionExcel = ''
                csvReportFolder = ''
                batchDict = {}
                try:
                    # grab the text of the polarion excel filepath and the csv report folder filepath from the batch table
                    polarionExcel = batchModel.item(row, polarionExcelCol).text()
                    csvReportFolder = batchModel.item(row, csvReportFolderCol).text()
                except:
                    pass
                try:
                    # place the polarionExcel text and csvReportFolder text into a dictionary of two
                    batchDict = {
                        'Polarion Excel File': polarionExcel,
                        'CSV Report Folder': csvReportFolder
                    }
                    # add the dictionary to the list
                    batchList.append(batchDict)
                except:
                    pass

            self.profileDict = {
                'Profile': {
                    '@version': '1.0',
                    # 'PolarionExcel': self.getCurrentExcelPath(),
                    # 'TestCaseExcel': self.getCurrentExcelPath(),
                    # 'CallFunctionFolder': self.callFunctionEdit.text(),
                    # 'CSVReportFolder': self.csvReportEdit.text(),
                    'VariablePoolPath': self.variablePoolEdit.text(),
                    'UpdateVariablePool': self.updateVariablePool,
                    'Version': {'@include': self.versionCheckBox.isChecked()},
                    'CallFunctionDebug': {'@enable': self.callFunctionDebugCheckbox.isChecked()},
                    'Log': {
                        '@mode': logMode,
                        'Signal': signalList
                    },
                    'DTC': {
                        '@enable': self.dtcExCheckBox.isChecked(),
                        'Except': dtcExList
                    },
                    'Versions': versionDict,
                    # 'UseGuiRunList': self.useGuiRunListCheckbox.isChecked(),
                    'PauseEnable': self.pauseCheckbox.isChecked(),
                    'PauseStep': str(self.pauseStepSpinbox.value()),
                    'PrintStepsInOutput': self.printStepsInOutputCheckbox.isChecked(),
                    'MaxPrintsInOutput': self.maxPrintsInOutputSpinbox.value(),
                    'IncludeTitle': self.titleCheckBox.isChecked(),
                    'IncludeDescription': self.descriptionCheckBox.isChecked(),
                    # 'Batch': batchDict
                    'Batch': batchList  # 'Batch' : [ {'Polarion Excel File': ..., 'CSV Report Folder': ...}, {...}]
                }
            }
        except:
            print(traceback.format_exc())

    def dtcTableModelItemChanged(self, item):
        """Handler for dtc model changed"""
        try:
            model = self.dtcTableModel
            header = self.dtcTableHeader
            try:
                dtcExList = self.profileDict['Profile']['DTC']['Except']
                row = item.row()
                column = item.column()
                colName = header[column]

                if column > 0:
                    if column == header.index('OBD'):
                        try:
                            dtcExList[row][colName] = model.item(row,column).checkState()
                        except:
                            dtcExList[row][colName] = False
                    else:
                        # print(row, colName, item.text())
                        try:
                            dtcExList[row][colName] = item.text()
                        except:
                            pass
            except KeyError:
                print('KeyError for dtcTableModelItemChanged, profileDict')
        except:
            print(traceback.format_exc())

    def addSignal(self):
        """Add signal to signal list"""
        try:
            signal = self.addSignalEdit.text()
            try:
                variablePoolKeys = list(self.variablePoolDict)

                try:
                    foundInVp = variablePoolKeys.index(signal) >= 0
                except:
                    foundInVp = False

                foundInModel = self.addSignalModel.findItems(signal)

                if foundInModel:
                    self.statusbar.showMessage('Duplicate signal. Try again')

                if foundInVp and not foundInModel:
                    self.addSignalModel.appendRow(QStandardItem(self.addSignalEdit.text()))
                    self.addSignalEdit.clear()
                    self.addSignalEdit.setFocus()
                    self.statusbar.showMessage('Signal added...OK!')
            except:
                self.statusbar.showMessage('Signal not found in variable pool')
        except:
            print(traceback.format_exc())

    def removeSignal(self):
        """Remove signal from signal list"""
        try:
            self.addSignalModel.removeRow(self.addSignalListView.currentIndex().row())
        except:
            print(traceback.format_exc())

    def addDtcEx2(self):
        """Add DTC exception to table"""
        try:
            hexCode = self.addDtcExEdit.text()
            model = self.dtcTableModel
            hexCodeItem = QStandardItem(hexCode)
            descItem = QStandardItem('')
            ecuItem = QStandardItem('')
            obdItem = QStandardItem(False)
            obdItem.setCheckable(True)

            model.appendRow([hexCodeItem, descItem, ecuItem, obdItem])
            self.addDtcExEdit.clear()
            self.addDtcExEdit.setFocus()
            self.statusbar.showMessage('DTC exception added...OK!')

            self.unsavedChanges = True
        except:
            print(traceback.format_exc())

    def removeDtcEx2(self):
        """Remove DTC exception from table"""
        try:
            model = self.dtcTableModel
            selectedRows = self.dtcTableView.selectionModel().selectedRows()

            if model.rowCount() > 0 and len(selectedRows) > 0:
                selectedItems = [model.item(x.row(), 0) for x in selectedRows]

                for item in selectedItems:
                    model.removeRow(item.row())

            self.unsavedChanges = True
        except:
            print(traceback.format_exc())

    def saveSettings(self):
        """Save settings"""
        try:
            self.settings = QSettings('Karma', 'AutomationGUI')
            self.settings.setValue('AutoRun', self.autorunCheckBox.isChecked())
            self.settings.setValue('AutoRunTimer', int(self.autorunSpinBox.value()))
            self.settings.setValue('LastProfile', self.profileFile)
            if not self.isMaximized():
                self.settings.setValue('Width', self.width())
                self.settings.setValue('Height', self.height())
            self.settings.setValue('PolarionUsername', self.polarionUsername)
            self.settings.setValue('PolarionLeftTableWidth', self.polarionTableView.width())

            del self.settings
        except:
            print(traceback.format_exc())

    def loadProfile(self):
        """Load profile"""
        try:
            # self.newProfile()
            try:
                with open(self.profileFile) as f:
                    self.profileDict = json.load(f)

                    profileValid = True

                    try:
                        profileversion = self.profileDict['Profile']['@version']
                        if profileversion == '1.0':
                            profileValid &= True
                        else:
                            debugPrint('Invalid profile version')
                    except:
                        debugPrint('No profile version defined')

                    varpoolfile = ''
                    try:
                        varpoolfile = self.profileDict['Profile']['VariablePoolPath']
                        if not os.path.exists(str(varpoolfile)):
                            self.statusbar.showMessage('Variable pool file not found')
                        profileValid &= True
                    except:
                        debugPrint('No variable pool file defined in profile')
                        self.statusbar.showMessage('No variable pool file found')

                    if profileValid:
                        # updates the gui after loading a valid profile
                        self.defaultProfile = False
                        self.updateGuiFromProfile()
                        if os.path.exists(str(varpoolfile)):
                            self.loadVariablePool()
                        self.setTitle(self.profileFile)
                        self.statusbar.showMessage('Profile loaded OK!')
                    else:
                        self.statusbar.showMessage('Invalid profile detected')
            except IOError:
                self.statusbar.showMessage('Profile does not exist. Please load profile')
            except ValueError:
                self.statusbar.showMessage('No valid profile found')
        except:
            print(traceback.format_exc())

    def checkFolderExist(self, path):
        """Check if folder exists"""
        try:
            basename = os.path.basename(str(path))
            if not os.path.exists(path):
                msgReply = QMessageBox.question(
                    self,
                    'Create Folder',
                    '\'' + basename + '\'' + ' folder was not found. Would you like to create it?',
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )

                if msgReply == QMessageBox.Yes:
                    os.mkdir(path)
        except:
            print(traceback.format_exc())

    def checkVarPoolExist(self, path):
        """Check if variable pool file exists"""
        try:
            if path is not None:
                basename = os.path.basename(str(path))
                if not os.path.exists(path):
                    msgReply = QMessageBox.question(
                        self,
                        'Varpool Not Found',
                        'Varpool {} does not exist. Browse for one?'.format(basename),
                        # '\'' + basename + '\'' + ' file was not found. Would you like to browse for one?',
                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                    )

                    if msgReply == QMessageBox.Yes:
                        self.browseVariablePool()
        except:
            print(traceback.format_exc())

    def getLogDict(self):
        """Enumerate the latest log files and extract results in new thread"""
        class getLogDictThread(QThread):
            showLoadingBarSignal = pyqtSignal()
            hideLoadingBarSignal = pyqtSignal()
            appendMessageSignal = pyqtSignal('PyQt_PyObject')

            def __init__(self):
                QThread.__init__(self)

            def run(self):
                try:
                    self.showLoadingBarSignal.emit()
                    self.appendMessageSignal.emit('Loading log results')

                    # enumerate the log folder and determine the latest log files
                    for fileName in os.listdir(self.csvReportFolder):
                        if fileName.endswith('.csv'):
                            filePath = os.path.join(self.csvReportFolder, fileName)
                            lastModTime = datetime.fromtimestamp(os.path.getmtime(filePath))
                            testCaseName = fileName.split('_')[2]

                            passList = []
                            testCaseVerdict = ''
                            testComment = ''
                            actualResult = []
                            length = 0

                            testCase = testCaseName.rstrip('.csv')

                            try:
                                # if testcase exist in dict, then we update it with latest csv filepath and modified date
                                if lastModTime > self.logDict[testCase]['lastModTime']:
                                    self.logDict[testCase]['filePath'] = filePath
                                    self.logDict[testCase]['lastModTime'] = lastModTime
                            except KeyError:
                                # create key-value pairs if test case does not exist in dict
                                self.logDict[testCase] = {'filePath': filePath,
                                                          'lastModTime': lastModTime,
                                                          'testCaseVerdict': testCaseVerdict,
                                                          'testComment': testComment,
                                                          'actualResult': actualResult,
                                                          'passList': passList,
                                                          'length': length}

                    # grab the data from all latest log files
                    for t in self.logDict:
                        passList = []
                        testCaseVerdict = ''
                        testComment = ''
                        actualResult = []
                        statusMessage = []

                        filePath = None

                        try:
                            filePath = self.logDict[t]['filePath']
                        except KeyError:
                            pass

                        if filePath is not None:
                            with open(filePath) as csvFile:
                                csvReader = csv.DictReader(csvFile)
                                for row in csvReader:
                                    try:
                                        passList.append(row['Step Verdict'])
                                    except KeyError:
                                        pass
                                    try:
                                        actualResult.append(row['Actual Value'])
                                    except KeyError:
                                        pass
                                    try:
                                        statusMessage.append(row['Step Message'])
                                    except KeyError:
                                        pass
                                    if len(row['Test Case Verdict']) > 0:
                                        testCaseVerdict = row['Test Case Verdict']
                                    if len(row['Test Comment']) > 0:
                                        testComment = row['Test Comment']

                            self.logDict[t]['testCaseVerdict'] = testCaseVerdict
                            self.logDict[t]['testComment'] = testComment
                            self.logDict[t]['actualResult'] = actualResult
                            self.logDict[t]['passList'] = passList
                            self.logDict[t]['length'] = len(passList)
                    self.appendMessageSignal.emit('Log results loaded')
                except:
                    print(traceback.format_exc())
                finally:
                    self.hideLoadingBarSignal.emit()

        try:
            self.logDict.clear()
            self.getLogDictThread = getLogDictThread()

            myThread = self.getLogDictThread
            myThread.logDict = self.logDict
            myThread.showLoadingBarSignal.connect(self.showLoadingBar)
            myThread.hideLoadingBarSignal.connect(self.hideLoadingBar)
            myThread.hideLoadingBarSignal.connect(self.saveLogResults)
            myThread.appendMessageSignal.connect(self.appendPolarionLog)
            myThread.csvReportFolder = self.getCurrentCsvReportFolder()
            myThread.start()
        except:
            print(traceback.format_exc())

    def savePolarionExcel(self):
        """Save to Polarion excel file in new thread"""
        class savePolarionExcelThread(QThread):
            showLoadingBarSignal = pyqtSignal()
            hideLoadingBarSignal = pyqtSignal()
            appendMessageSignal = pyqtSignal('PyQt_PyObject')
            updatePolarionRevisionSignal = pyqtSignal()

            def __init__(self):
                QThread.__init__(self)

            def run(self):
                try:
                    if len(self.filePath) > 0:
                        self.showLoadingBarSignal.emit()
                        try:
                            # udpate polarion excel file

                            if self.polarionWb:
                                self.polarionWb.close()

                            self.polarionWb = load_workbook(filename=self.polarionExcelPath)
                            self.polarionWs = self.polarionWb.get_sheet_by_name('Sheet1')
                            # print(self.polarionWs.max_row)

                            # self.updatePolarionRevisionSignal.emit()

                            count = 0
                            for t in self.polarionDict:
                                try:
                                    if self.polarionDict[t]['length'] == self.logDict[t]['length']:
                                        if self.polarionDict[t]['testCaseVerdict'] in ['Passed', 'Deferred', 'Error']:
                                            # print(t, 'Match')
                                            startRow = self.polarionDict[t]['startRow']
                                            endRow = self.polarionDict[t]['endRow']

                                            actualResultRowStart = 'M{}'.format(startRow)
                                            actualResultRowEnd = 'M{}'.format(endRow)

                                            actualResultRows = self.polarionWs[actualResultRowStart:actualResultRowEnd]

                                            for row, actual in zip(actualResultRows, self.logDict[t]['actualResult']):
                                                row[0].value = str(actual)

                                            stepVerdictRowStart = 'N{}'.format(startRow)
                                            stepVerdictRowEnd = 'N{}'.format(endRow)

                                            stepVerdictRows = self.polarionWs[stepVerdictRowStart:stepVerdictRowEnd]
                                            for row, verdict in zip(stepVerdictRows, self.logDict[t]['passList']):
                                                row[0].value = verdict

                                            testCaseVerdictCellIndex = 'O{}'.format(startRow)
                                            self.polarionWs[testCaseVerdictCellIndex].value = self.logDict[t][
                                                'testCaseVerdict']

                                            testCaseCommentIndex = 'P{}'.format(startRow)
                                            self.polarionWs[testCaseCommentIndex].value = self.logDict[t]['testComment']

                                            count += 1
                                except KeyError:
                                    pass

                            self.polarionWb.save(self.filePath)
                            self.appendMessageSignal.emit('Save successful')
                        except IOError as error:
                            if str(error).__contains__('Permission denied'):
                                self.appendMessageSignal.emit('Permission Denied. Unable to save excel file')
                except:
                    print(traceback.format_exc())
                finally:
                    self.hideLoadingBarSignal.emit()

        try:

            folder = os.path.dirname(self.getCurrentExcelPath())
            filePath, fileType = QFileDialog.getSaveFileName(
                self,
                "Save Polarion Excel File",
                folder,
                'XLSX Files (*.xlsx);;All Files (*)'
            )

            self.savePolarionExcelThreadObject = savePolarionExcelThread()
            myThread = self.savePolarionExcelThreadObject
            myThread.filePath = filePath
            myThread.polarionExcelPath = self.getCurrentExcelPath()
            myThread.polarionWb = self.polarionWb
            myThread.polarionWs = self.polarionWs
            myThread.polarionDict = self.polarionDict
            myThread.logDict = self.logDict
            myThread.showLoadingBarSignal.connect(self.showLoadingBar)
            myThread.hideLoadingBarSignal.connect(self.hideLoadingBar)
            myThread.appendMessageSignal.connect(self.appendPolarionLog)
            myThread.updatePolarionRevisionSignal.connect(self.udpatePolarionRevision)
            myThread.start()
        except:
            print(traceback.format_exc())

    def readPolarionExcel(self):
        """Read and process the polarion excel file in new thread"""
        class readPolarionExcelThread(QThread):
            showLoadingBarSignal = pyqtSignal()
            hideLoadingBarSignal = pyqtSignal()
            finishedSignal = pyqtSignal('PyQt_PyObject')
            updatePolarionRevisionSignal = pyqtSignal('PyQt_PyObject')
            appendMessageSignal = pyqtSignal('PyQt_PyObject')

            def __init__(self):
                QThread.__init__(self)

            def run(self):
                try:
                    self.appendMessageSignal.emit('Loading Polarion Excel')
                    self.showLoadingBarSignal.emit()

                    # open polarion excel file
                    book = xlrd.open_workbook(filename=self.polarionExcel)
                    sheet = book.sheet_by_name('Sheet1')


                    testStepList = []

                    polarionDict = self.polarionDict

                    colNum = 18

                    startRow = 0
                    endRow = 0
                    testCaseIdCol = 0
                    titleCol = 1
                    descCol = 2
                    stepCol = 3
                    waitCol = 10

                    if sheet.nrows > 1:
                        for i in range(1, sheet.nrows + 1):
                            try:
                                row = sheet.row_values(i, 0, colNum)
                            except IndexError:
                                pass

                            testCaseStep = row[stepCol]
                            if testCaseStep in ['1', '1.0', 1, 1.0] or i == sheet.nrows:
                                if len(testStepList) > 0:
                                    endRow = i
                                    length = len(testStepList)
                                    startRow = endRow - length + 1
                                    title = testStepList[0][titleCol]
                                    desc = testStepList[0][descCol]

                                    def conv2float(x):
                                        try:
                                            return float(x)
                                        except:
                                            return 0.0

                                    totalWaitTime = round(sum([conv2float(x[waitCol]) for x in testStepList]), 0)

                                    testcase = testStepList[0][testCaseIdCol]

                                    if testcase not in polarionDict:
                                        polarionDict[testcase] = {
                                            'title': None,
                                            'desc': None,
                                            'steps': None,
                                            'startRow': None,
                                            'endRow': None,
                                            'length': None,
                                            'totalWaitTime': None,
                                        }

                                    polarionDict[testcase]['title'] = title
                                    polarionDict[testcase]['desc'] = desc
                                    polarionDict[testcase]['steps'] = copy.copy(testStepList)
                                    polarionDict[testcase]['startRow'] = startRow
                                    polarionDict[testcase]['endRow'] = endRow
                                    polarionDict[testcase]['length'] = length
                                    polarionDict[testcase]['totalWaitTime'] = totalWaitTime

                                    testStepList = []
                            testStepList.append(row)


                        # read revision number
                        revisionSheet = book.sheet_by_name('_polarion')

                        for i in range(0, revisionSheet.nrows):
                            row = revisionSheet.row_values(i, 0, 1)
                            if row[0] == 'testRunRevision':
                                revisionRow = i
                                revisionNumber = revisionSheet.cell_value(revisionRow, 1)
                                self.updatePolarionRevisionSignal.emit(str(revisionNumber))
                                break

                    self.appendMessageSignal.emit('Polarion Excel loaded')
                    self.finishedSignal.emit(polarionDict)
                except:
                    print(traceback.format_exc())
                finally:
                    self.hideLoadingBarSignal.emit()

        try:

            self.polarionReadExcelThread = readPolarionExcelThread()
            myThread = self.polarionReadExcelThread

            # clear current polarion table view
            self.polarionTableViewModel.clear()
            self.polarionDict.clear()

            # try to load existing json file for selected polarion excel
            self.loadPolarionJson()

            myThread.polarionDict = self.polarionDict

            # bring in polarion excel selected from dropbox
            myThread.polarionExcel = self.getCurrentExcelPath()

            myThread.appendMessageSignal.connect(self.appendPolarionLog)
            myThread.updatePolarionRevisionSignal.connect(self.polarionRevisionLineEdit.setText)
            myThread.finishedSignal.connect(self.polarionReadExcelFinished)
            myThread.showLoadingBarSignal.connect(self.showLoadingBar)
            myThread.hideLoadingBarSignal.connect(self.hideLoadingBar)
            myThread.start()
        except:
            print(traceback.format_exc())

    def loadPolarionWorkBook(self):
        """Load Polarion excel workbook"""
        try:
            polarionExcel = self.getCurrentExcelPath()
            self.polarionWb = load_workbook(filename=polarionExcel, read_only=True)
            self.polarionWs = self.polarionWb.active
            if self.polarionReadExcelThread is not None:
                self.polarionReadExcelThread.polarionWb = self.polarionWb
                self.polarionReadExcelThread.polarionWs = self.polarionWs
        except:
            print(traceback.format_exc())

    def updateDtcTableModel(self):
        """Update DTC table model"""
        try:
            model = self.dtcTableModel
            view = self.dtcTableView
            header = self.dtcTableHeader
            model.clear()
            model.setHorizontalHeaderLabels(header)

            hexCodeCol = header.index('Hex Code')
            descCol = header.index('Description')
            moduleCol = header.index('Module')
            obdCol = header.index('OBD')

            try:
                dtcExList = self.profileDict['Profile']['DTC']['Except']
                for row, dtc in enumerate(dtcExList):
                    hexCodeItem = QStandardItem(dtc['Hex Code'])
                    descItem = QStandardItem(dtc['Description'])
                    moduleItem = QStandardItem(dtc['Module'])
                    try:
                        isOBD = dtc['OBD']
                    except:
                        isOBD = False

                    obdItem = QStandardItem(isOBD)
                    obdItem.setCheckable(True)
                    obdItem.setCheckState(Qt.Checked if isOBD else Qt.Unchecked)

                    model.setItem(row, hexCodeCol, hexCodeItem)
                    model.setItem(row, descCol, descItem)
                    model.setItem(row, moduleCol, moduleItem)
                    model.setItem(row, obdCol, obdItem)
            except:
                print('KeyError for updateGuiFromProfileDict, dtcExList2')

            view.resizeColumnsToContents()
            model.itemChanged.connect(self.setUnsavedChanges)
        except:
            print(traceback.format_exc())

    def updateVersionTableModel(self):
        """Update version table model"""
        try:
            header = self.versionTableHeader
            view = self.versionTableView
            model = self.versionTableModel
            model.clear()
            model.setHorizontalHeaderLabels(header)

            try:
                versionDict = self.profileDict['Profile']['Versions']
                ecuCol = header.index('ECU')
                swVersionCol = header.index('Software')
                hwVersionCol = header.index('Hardware')

                for ecu in versionDict:
                    ecuItem = QStandardItem(ecu)
                    ecuItem.setEditable(False)

                    try:
                        swVersionItem = QStandardItem(versionDict[ecu]['Software'])
                    except:
                        swVersionItem = QStandardItem('')

                    try:
                        hwVersionItem = QStandardItem(versionDict[ecu]['Hardware'])
                    except:
                        hwVersionItem = QStandardItem('')

                    model.appendRow([ecuItem, swVersionItem, hwVersionItem])



                view.resizeColumnsToContents()
                view.setAlternatingRowColors(True)
            except KeyError:
                print('There was an error with updating version table model')

            self.versionTableModel.itemChanged.connect(self.setUnsavedChanges)
        except:
            print(traceback.format_exc())

    def versionAddRow(self):
        """Add version of a module"""
        try:
            model = self.versionTableModel
            model.appendRow(QStandardItem(''))
            self.unsavedChanges = True
        except:
            print(traceback.format_exc())

    def versionRemoveRow(self):
        """Remove version for a module"""
        try:
            model = self.versionTableModel
            selectedRows = self.versionTableView.selectionModel().selectedRows()

            if model.rowCount() > 0 and len(selectedRows) > 0:
                selectedItems = [model.item(x.row(), 0) for x in selectedRows]

                for item in selectedItems:
                    model.removeRow(item.row())

            self.unsavedChanges = True
        except:
            print(traceback.format_exc())

    def versionTableViewItemChanged(self, item):
        """Handler for version table model changed"""
        try:
            model = self.versionTableModel
            header = self.versionTableHeader
            row = item.row()
            column = item.column()
            versionDict = self.profileDict['Profile']['Versions']
            ecuCol = header.index('ECU')
            ecu = model.item(row, ecuCol).text()
            colName = header[column]
            if column > 0 and len(ecu) > 0:
                try:
                    versionDict[ecu][colName] = item.text()
                except KeyError:
                    versionDict[ecu] = {}
                    versionDict[ecu][colName] = item.text()
        except:
            print(traceback.format_exc())

    def updateBatchTableModel(self):
        """Update batch table model based on existing profile dict"""
        try:
            header = self.batchTableHeader
            view = self.batchTableView
            model = self.batchTableModel
            model.clear()
            model.setHorizontalHeaderLabels(header)

            try:
                # batchDict = self.profileDict['Profile']['Batch']
                # polarionExcelCol = header.index('Polarion Excel File')
                # csvReportCol = header.index('CSV Report Folder')
                #
                #
                # for polarionExcel in batchDict:
                #     polarionExcelItem = QStandardItem(polarionExcel)
                #
                #     try:
                #         csvReportItem = QStandardItem(batchDict[polarionExcel]['CSV Report Folder'])
                #     except:
                #         csvReportItem = QStandardItem('')
                #
                #     model.appendRow([polarionExcelItem, csvReportItem])

                batchList = self.profileDict['Profile']['Batch']
                polarionExcelCol = header.index('Polarion Excel File')
                csvReportCol = header.index('CSV Report Folder')

                for batchDict in batchList:
                    # get polarion excel file and csv report folder from each batch dictionary of the list and add to table
                    try:
                        polarionExcelItem = QStandardItem(batchDict['Polarion Excel File'])
                    except:
                        polarionExcelItem = QStandardItem('')
                    try:
                        csvReportFolderItem = QStandardItem(batchDict['CSV Report Folder'])
                    except:
                        csvReportFolderItem = QStandardItem('')
                    model.appendRow([polarionExcelItem, csvReportFolderItem])


                view.resizeColumnsToContents()
                view.setAlternatingRowColors(True)
            except KeyError:
                print('There was an error with updating batch table model')

            self.batchTableModel.itemChanged.connect(self.setUnsavedChanges)
        except:
            print(traceback.format_exc())

    def batchAddRow(self):
        """Add test batch to running queue"""
        try:
            model = self.batchTableModel

            # prompt user to select an excel file to add to batch list
            filePath = self.browseFile('', 'Select Polarion TestCase Excel File', 'Excel Files (*.xlsx)')
            if filePath == '':
                return

            polarionExcelItem = QStandardItem(filePath)

            # create csv report folder based
            testCaseExcelPath = Path(filePath)
            dirName = os.path.dirname(str(testCaseExcelPath))
            newCsvReportFolder = os.path.join(dirName, 'Logs')
            csvReportItem = None
            if os.path.exists(dirName):
                msgReply = QMessageBox.question(
                    self,
                    'CSV Report Folder',
                    'Use \'{}\' to store reports?'.format(newCsvReportFolder),
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )

                if msgReply == QMessageBox.Yes:
                    csvReportItem = QStandardItem(str(newCsvReportFolder))
                    if not os.path.exists(newCsvReportFolder):
                        try:
                            os.mkdir(newCsvReportFolder)
                        except:
                            pass

            # add the row containing the excel file path and csv report folder
            model.appendRow([polarionExcelItem,csvReportItem])
            self.batchTableView.resizeColumnsToContents()
            self.unsavedChanges = True

        except:
            print(traceback.format_exc())

    def batchRemoveRow(self):
        """ Remove selected rows from batch table queue """
        """ """
        try:
            model = self.batchTableModel
            view = self.batchTableView
            header = self.batchTableHeader
            selectedIndexes = view.selectedIndexes()

            if model.rowCount() > 0 and len(selectedIndexes) > 0:
                # make list of items to remove
                selectedItems = [model.item(x.row(), 0) for x in selectedIndexes]

                for item in selectedItems:
                    model.removeRow(item.row())

            self.setUnsavedChanges()
        except:
            print(traceback.format_exc())

    def batchClearAll(self):
        """ Clear all items in batch table queue """
        try:
            model = self.batchTableModel
            if model.rowCount() > 0:
                model.removeRows(0, model.rowCount())

            self.setUnsavedChanges()

        except:
            print(traceback.format_exc())

    def polarionReadExcelFinished(self, polarionDict):
        """Slot for reading excel finished"""
        try:
            self.polarionDict = polarionDict
            self.savePolarionDict()
            # self.appendPolarionLog(msg)
            self.updatePolarionTableModel()
        except:
            print(traceback.format_exc())

    def updatePolarionTableModel(self):
        """Update the model for Polarion table view"""
        try:
            model = self.polarionTableViewModel
            header = self.polarionTableHeader
            model.setHorizontalHeaderLabels(header)

            testCaseCol = header.index('TestCase')
            stepsTotalCol = header.index('Steps')
            totalWaitTimeCol = header.index('Total Wait Time')
            runTestCol = header.index('Run Test')
            stepVerdictCol = header.index('TestCase Verdict')
            modifiedCol = header.index('Modified')
            commentsCol = header.index('Comments')
            hyperlinksCol = header.index('Hyperlinks')

            polarionDict = self.polarionDict

            for row, d in enumerate(polarionDict):
                testCaseItem = QStandardItem()
                testCaseItem.setCheckable(True)
                testCaseItem.setEditable(False)
                try:
                    testCaseItem.setText(d)
                except:
                    pass
                model.setItem(row, testCaseCol, testCaseItem)


                lengthItem = QStandardItem()
                lengthItem.setEditable(False)
                try:
                    lengthItem.setData(polarionDict[d]['length'], Qt.DisplayRole)
                except KeyError:
                    pass
                model.setItem(row, stepsTotalCol, lengthItem)


                timeItem = QStandardItem()
                timeItem.setEditable(False)
                try:
                    timeItem.setData(polarionDict[d]['totalWaitTime'], Qt.DisplayRole)
                except KeyError:
                    pass
                model.setItem(row, totalWaitTimeCol, timeItem)


                testCaseVerdictItem = QStandardItem()
                testCaseVerdictItem.setEditable(False)
                try:
                    testCaseVerdict = polarionDict[d]['testCaseVerdict']
                    testCaseVerdictItem.setText(testCaseVerdict)
                    if testCaseVerdict in ['Deferred', 'Error', 'Not Passed']:
                        testCaseVerdictItem.setIcon(QIcon(':/icon/failed'))
                    elif testCaseVerdict in ['Passed']:
                        testCaseVerdictItem.setIcon(QIcon(':/icon/passed'))
                except KeyError:
                    pass
                model.setItem(row, stepVerdictCol, testCaseVerdictItem)


                modifiedItem = QStandardItem()
                modifiedItem.setCheckable(True)
                modifiedItem.setEditable(False)
                try:
                    modified = polarionDict[d]['modified']
                    modifiedItem.setText(str(modified))
                    modifiedItem.setCheckState(Qt.Checked if modified else Qt.Unchecked)
                except:
                    pass
                model.setItem(row, modifiedCol, modifiedItem)


                commentsItem = QStandardItem()
                try:
                    comments = polarionDict[d]['comments']
                    commentsItem.setText(comments)
                except:
                    pass
                model.setItem(row, commentsCol, commentsItem)


                runTestItem = QStandardItem()
                runTestItem.setEditable(False)
                runTestItem.setCheckable(True)
                try:
                    runTest = polarionDict[d]['run']
                    runTestItem.setText(str(runTest))
                    runTestItem.setCheckState(Qt.Checked if runTest else Qt.Unchecked)
                except KeyError as err:
                    runTestItem.setText('False')
                    runTestItem.setCheckState(Qt.Unchecked)
                model.setItem(row, runTestCol, runTestItem)


                hyperlinksItem = QStandardItem()
                hyperlinksItem.setEditable(False)
                try:
                    hyperlinks = polarionDict[d]['hyperlinks']
                    hyperlinksItem.setText(hyperlinks)
                except:
                    pass
                model.setItem(row, hyperlinksCol, hyperlinksItem)

            self.polarionTableView.resizeColumnsToContents()
            self.polarionTableView.setColumnWidth(testCaseCol, 150)

            self.polarionTableView.setSortingEnabled(True)
            # self.polarionTableView.sortByColumn(testCaseCol, Qt.AscendingOrder)
            # self.polarionTableView.setAlternatingRowColors(True)

            logDict = self.logDict

            allTestCaseVerdicts = [logDict[x]['testCaseVerdict'] for x in logDict]

            passedTotal = allTestCaseVerdicts.count('Passed')
            deferredTotal = allTestCaseVerdicts.count('Deferred')
            errorTotal = allTestCaseVerdicts.count('Error')
            total = passedTotal + deferredTotal + errorTotal

            self.polarionOverViewLineEdit.setText(
                'Total Logs Found: {}    Passed: {}    Deferred: {}    Error: {}'.format(total, passedTotal, deferredTotal,
                                                                                         errorTotal))
            self.hideLoadingBar()

            self.polarionTableViewModel.itemChanged.connect(self.polarionTableViewModelItemChanged)
            self.polarionTableViewModel.itemChanged.connect(self.setUnsavedChanges)

            self.statusbar.showMessage('Polarion table updated')
        except:
            print(traceback.format_exc())

    def updatePolarionVerdicts(self, testCaseVerdict=''):
        """Update Polarion dict with verdicts from logs"""
        try:
            self.getLogDict()
            self.getLogDictThread.hideLoadingBarSignal.connect(lambda: self.updatePolarionDict(testCaseVerdict))
        except:
            print(traceback.format_exc())

    def updatePolarionDict(self, testCaseVerdict):
        """Update Polarion dict with specific verdicts"""
        try:
            polarionDict = self.polarionDict
            logDict = self.logDict
            for t in polarionDict:
                polarionDict[t]['testCaseVerdict'] = ''

                try:
                    if logDict[t]['testCaseVerdict'] == testCaseVerdict or testCaseVerdict == 'All':
                        polarionDict[t]['testCaseVerdict'] = logDict[t]['testCaseVerdict']
                    if polarionDict[t]['length'] != logDict[t]['length']:
                        polarionDict[t]['comments'] = 'Log steps mistmatch'
                        # polarionDict[t]['modified'] = True

                except KeyError:
                    self.polarionDict[t]['testCaseVerdict'] = 'No log found'

            self.updatePolarionTableModel()
            self.appendPolarionLog('Updated excel file with {} verdicts'.format(testCaseVerdict))
            self.getLogDictThread.disconnect()
        except:
            print(traceback.format_exc())

    def udpatePolarionRevision(self):
        """Update revision number in Polarion excel"""
        try:
            polarionWs = self.polarionWb['_polarion']
            polarionColA = [x.value for x in polarionWs['A']]
            revisionRow = polarionColA.index('testRunRevision') + 1
            polarionWs.cell(row=revisionRow, column=2).value = self.polarionRevisionLineEdit.text()

            self.appendPolarionLog('Updated Polarion revision number')
        except:
            print(traceback.format_exc())

    def updateHyperlinks(self):
        """Update hyperlinks in Polarion table view"""
        class updateHyperlinksThread(QThread):
            showLoadingBarSignal = pyqtSignal()
            hideLoadingBarSignal = pyqtSignal()
            finishedSignal = pyqtSignal('PyQt_PyObject')
            appendMessageSignal = pyqtSignal('PyQt_PyObject')

            def __init__(self):
                QThread.__init__(self)

            class TransportSubclass(zeep.transports.Transport):
                def __init__(self, *args, **kwargs):
                    super(self.__class__, self).__init__(*args, **kwargs)
                    self.last_response = None

                def post(self, *args, **kwargs):
                    self.last_response = super(self.__class__, self).post(*args, **kwargs)
                    return self.last_response

            def run(self):
                # login session and get session id
                session = requests.Session()
                transport = self.TransportSubclass(session=session)
                loginClientWsdl = 'http://polarion.karmaautomotive.com/polarion/ws/services/SessionWebService?wsdl'
                try:
                    loginClient = zeep.Client(wsdl=loginClientWsdl, transport=transport)

                    try:
                        loginClient.service.logIn(self.username, self.password)

                        self.showLoadingBarSignal.emit()
                        # grab repsonse header and parse for session id
                        root = etree.XML(transport.last_response.content)
                        sessions = root.xpath('//ns1:sessionID', namespaces={'ns1': 'http://ws.polarion.com/session'})
                        session_id = sessions[0]

                        # use the new session id for transports
                        transport = zeep.transports.Transport(session=session)

                        # use session id for web service
                        testWsdl = 'http://polarion.karmaautomotive.com/polarion/ws/services/TestManagementWebService?wsdl'
                        testClient = zeep.Client(testWsdl, transport=transport)
                        testClient._default_soapheaders = [session_id]

                        # use session id for web service
                        trackerClientWsdl = 'http://polarion.karmaautomotive.com/polarion/ws/services/TrackerWebService?wsdl'
                        trackerClient = zeep.Client(trackerClientWsdl, transport=transport)
                        trackerClient._default_soapheaders = [session_id]

                        queryString = ' '.join(self.polarionDict.keys())

                        # query = trackerClient.service.queryWorkItems(
                        #     query='testtype.KEY:manual AND status:(reviewed review) AND ecu.KEY:VCMB', sort='hyperlinks',
                        #     fields=['id', 'hyperlinks'])

                        query = trackerClient.service.queryWorkItems(
                            query=queryString, sort='hyperlinks',
                            fields=['id', 'hyperlinks'])

                        loginClient.service.endSession()

                        hyperDict = {}
                        for q in query:
                            try:
                                hyperlinks = [x.uri for x in q.hyperlinks.Hyperlink]
                            except:
                                hyperlinks = ''
                            hyperDict[q.id] = {'hyperlinks': hyperlinks}

                        self.finishedSignal.emit(hyperDict)
                    except zeep.exceptions.Fault as error:
                        if error.message.__contains__('Authentication failed'):
                            self.appendMessageSignal.emit('Authentication failed. Invalid username or password')

                except requests.exceptions.HTTPError:
                    self.appendMessageSignal.emit('Unable to pull hyperlinks')
                    pass

                self.hideLoadingBarSignal.emit()

        try:
            self.getPolarionAccount()
            if self.polarionUsername != '' and self.polarionPassword != '':
                self.updateHyperlinksThreadObject = updateHyperlinksThread()
                myThread = self.updateHyperlinksThreadObject
                myThread.polarionDict = self.polarionDict
                myThread.appendMessageSignal.connect(self.appendPolarionLog)
                myThread.showLoadingBarSignal.connect(self.showLoadingBar)
                myThread.hideLoadingBarSignal.connect(self.hideLoadingBar)
                myThread.username = self.polarionUsername
                myThread.password = self.polarionPassword
                myThread.finishedSignal.connect(self.updatePolarionDictWithHyperlinks)
                myThread.start()
        except:
            print(traceback.format_exc())

    def updatePolarionSteps(self):
        """Update Polarion server with modified steps in new thread"""
        class updatePolarionStepsThread(QThread):
            finishedSignal = pyqtSignal('PyQt_PyObject')
            appendMessageSignal = pyqtSignal('PyQt_PyObject')
            showLoadingBarSignal = pyqtSignal()
            hideLoadingBarSignal = pyqtSignal()

            def __init__(self):
                QThread.__init__(self)

            class TransportSubclass(zeep.transports.Transport):
                def __init__(self, *args, **kwargs):
                    super(self.__class__, self).__init__(*args, **kwargs)
                    self.last_response = None

                def post(self, *args, **kwargs):
                    self.last_response = super(self.__class__, self).post(*args, **kwargs)
                    return self.last_response

            def run(self):
                self.appendMessageSignal.emit('Loading Polarion excel')

                book = xlrd.open_workbook(filename=self.polarionExcel)
                sheet = book.sheet_by_name('Sheet1')

                testStepList = []
                polarionDict = {}
                startRow = 0
                endRow = 0
                testCaseIdCol = 0
                titleCol = 1
                descCol = 2
                stepCol = 3
                waitCol = 10
                startCol = 0
                endCol = 12
                if sheet.nrows > 1:
                    for i in range(1, sheet.nrows + 1):
                        try:
                            row = sheet.row_values(i, startCol, endCol)
                        except IndexError:
                            pass

                        testCaseStep = row[stepCol]
                        if testCaseStep in ['1', '1.0', 1, 1.0] or i == sheet.nrows:
                            if len(testStepList) > 0:
                                endRow = i
                                length = len(testStepList)
                                startRow = endRow - length + 1
                                title = testStepList[0][titleCol]
                                desc = testStepList[0][descCol]

                                def conv2float(x):
                                    try:
                                        return float(x)
                                    except:
                                        return 0.0

                                totalWaitTime = round(sum([conv2float(x[waitCol]) for x in testStepList]), 1)

                                testStepListCompact = [x[4:] for x in testStepList]

                                polarionDict[testStepList[0][testCaseIdCol]] = {
                                    'title': title,
                                    'desc': desc,
                                    'steps': copy.copy(testStepList),
                                    'stepsCompact': copy.copy(testStepListCompact),
                                    'startRow': startRow,
                                    'endRow': endRow,
                                    'length': length,
                                    'testCaseVerdict': '',
                                    'hyperlinks': '',
                                    'run': 'No',
                                    'totalWaitTime': totalWaitTime
                                }
                                testStepList = []
                        testStepList.append(row)

                    # with open("C:\\Users\\pthil\\Documents\\TestRuns\\BL1015\\DMLR_Automated_V10_2019_7_29\\polarionDict.json", 'w') as f:
                    #     polarionJson = json.dumps(polarionDict, sort_keys=True, indent=4)
                    #     f.write(polarionJson)

                    # read revision number
                    revisionSheet = book.sheet_by_name('_polarion')

                    for i in range(0, revisionSheet.nrows):
                        row = revisionSheet.row_values(i, 0, 1)
                        if row[0] == 'testRunRevision':
                            revisionRow = i
                            revisionNumber = revisionSheet.cell_value(revisionRow, 1)
                            # self.updatePolarionRevisionSignal.emit(revisionNumber)
                            break

                self.appendMessageSignal.emit('Polarion excel loaded')

                if True:
                    try:
                        self.appendMessageSignal.emit('Connecting to Polarion server')

                        # login session and get session id
                        session = requests.Session()
                        transport = self.TransportSubclass(session=session)
                        loginClientWsdl = 'http://polarion.karmaautomotive.com/polarion/ws/services/SessionWebService?wsdl'
                        loginClient = zeep.Client(wsdl=loginClientWsdl, transport=transport)

                        try:
                            loginClient.service.logIn(self.username, self.password)
                            self.showLoadingBarSignal.emit()
                            # grab repsonse header and parse for session id
                            root = etree.XML(transport.last_response.content)
                            sessions = root.xpath('//ns1:sessionID', namespaces={'ns1': 'http://ws.polarion.com/session'})
                            session_id = sessions[0]

                            # use the new session id for transports
                            transport = zeep.transports.Transport(session=session)

                            # use session id for web service
                            testWsdl = 'http://polarion.karmaautomotive.com/polarion/ws/services/TestManagementWebService?wsdl'
                            testClient = zeep.Client(testWsdl, transport=transport)
                            testClient._default_soapheaders = [session_id]

                            # use session id for web service
                            trackerClientWsdl = 'http://polarion.karmaautomotive.com/polarion/ws/services/TrackerWebService?wsdl'
                            trackerClient = zeep.Client(trackerClientWsdl, transport=transport)
                            trackerClient._default_soapheaders = [session_id]

                            self.appendMessageSignal.emit('Connected to Polarion server')

                            # column indexes
                            columnIndexes = {
                                'Phase': 0,
                                'Action': 1,
                                'Description': 2,
                                'Variable': 3,
                                'Setting': 4,
                                'ExpectedResult': 5,
                                'Wait': 6,
                                'Remark': 7
                            }

                            # testStepCopy = None
                            self.appendMessageSignal.emit('Updating testcases')
                            for id in self.testCasesList:
                                # get workitem from polarion using the testcase id
                                query = trackerClient.service.queryWorkItems(query=id, sort='id', fields=['id', 'type'])

                                # do update if a query returns
                                testCaseUri = ''
                                if len(query) > 0:
                                    for q in query:
                                        if q['type']['id'] == 'testCase' and id in q['uri']:
                                            testCaseUri = q['uri']
                                            break

                                    # double check that id and test case uri matches
                                    if id in testCaseUri:
                                        # grab steps from query
                                        # if testStepCopy is None:
                                        querySteps = testClient.service.getTestSteps(testCaseUri)['steps']['TestStep']

                                        updatedQuerySteps = []

                                        if len(querySteps) > 0:
                                            # fill updated query steps, use the first step as a copy
                                            while len(updatedQuerySteps) < polarionDict[id]['length']:
                                                testStepCopy = copy.deepcopy(querySteps[0])
                                                updatedQuerySteps.append(testStepCopy)

                                            for i in range(0, polarionDict[id]['length']):
                                                for idx in columnIndexes.values():
                                                    updatedQuerySteps[i]['values']['Text'][idx]['content'] = \
                                                        polarionDict[id]['stepsCompact'][i][idx]

                                            # print(polarionDict[id]['stepsCompact'])
                                            # service to update test steps for specific testcase
                                            testClient.service.setTestSteps(testCaseUri, updatedQuerySteps)

                                            # print('Updated {}. Number of Steps: {} -> {}'.format(id, len(querySteps),
                                            #                                                       len(updatedQuerySteps)))
                                            self.appendMessageSignal.emit('Updated {}. Steps: {} -> {}'.format(id, len(querySteps), len(updatedQuerySteps)))
                                            # time.sleep(2.0)
                                else:
                                    self.appendMessageSignal.emit('{} was not found'.format(id))
                                    # print(id, 'not found')

                            # close session
                            loginClient.service.endSession()
                            # self.finishedSignal.emit('Updating Polarion finished')
                            self.appendMessageSignal.emit('Updating Polarion finished')
                            self.hideLoadingBarSignal.emit()
                        except zeep.exceptions.Fault as error:
                            self.hideLoadingBarSignal.emit()
                            if error.message.__contains__('Authentication failed'):
                                self.appendMessageSignal.emit('Authentication failed. Invalid username or password')
                    except:
                        self.hideLoadingBarSignal.emit()

        try:
            polarionExcel = self.browseFile(self.getCurrentExcelPath(), 'Open Polarion Excel file', 'XLSX Files (*.xlsx)')
            self.getPolarionAccount()

            if len(polarionExcel) > 0 and self.polarionUsername != '' and self.polarionPassword != '':
                self.udpatePolarionStepsThread = updatePolarionStepsThread()
                myThread = self.udpatePolarionStepsThread

                myThread.polarionExcel = polarionExcel
                myThread.polarionDict = self.polarionDict
                myThread.username = self.polarionUsername
                myThread.password = self.polarionPassword

                testCasesList = []
                model = self.polarionTableViewModel
                testCaseCol = self.polarionTableHeader.index('TestCase')
                modifiedCol = self.polarionTableHeader.index('Modified')

                for i in range(0, model.rowCount()):
                    testcaseItem = model.item(i, testCaseCol)
                    modifiedItem = model.item(i, modifiedCol)
                    if modifiedItem.checkState() == Qt.Checked:
                        testCasesList.append(testcaseItem.text())

                # print(testCasesList)

                myThread.testCasesList = testCasesList
                myThread.appendMessageSignal.connect(self.appendPolarionLog)
                myThread.showLoadingBarSignal.connect(self.showLoadingBar)
                myThread.hideLoadingBarSignal.connect(self.hideLoadingBar)
                # myThread.finishedSignal.connect(self.updatePolarionDictWithQuerySteps)
                myThread.start()
            else:
                self.appendPolarionLog('Update steps cancelled')
        except:
            print(traceback.format_exc())

    def getPolarionAccount(self):
        """Get Polarion account info from user"""
        try:
            username, okPressed1 = QInputDialog.getText(self, 'Enter Polarion Account', 'Username:',
                                                        QLineEdit.Normal, self.polarionUsername)
            password, okPressed2 = QInputDialog.getText(self, 'Enter Polarion Account', 'Password:',
                                                        QLineEdit.Password, self.polarionPassword)

            if okPressed1:
                self.polarionUsername = username
            if okPressed2:
                self.polarionPassword = password
        except:
            print(traceback.format_exc())

    def appendPolarionLog(self, msg):
        """Append to Polarion log"""
        try:
            self.polarionLogEdit.appendPlainText('{} - {}'.format(datetime.now().__str__(), msg))
        except:
            print(traceback.format_exc())

    def updatePolarionDictWithHyperlinks(self, hyperDict):
        """Update Polarion dict with hyperlinks"""
        try:
            for t in self.polarionDict:
                try:
                    hyperlinks = [x.strip() for x in hyperDict[t]['hyperlinks']]
                    hyperlinksStr = ', '.join(hyperlinks)

                    self.polarionDict[t]['hyperlinks'] = hyperlinksStr
                except KeyError:
                    pass
                except TypeError:
                    pass

            self.savePolarionDict()
            self.appendPolarionLog('Updated Polarion table with hyperlinks')
            self.updatePolarionTableModel()
        except:
            print(traceback.format_exc())

    def updateGuiFromProfile(self):
        """Update GUI from profile dictionary"""
        try:
            if 'Profile' in self.profileDict:
                profile = self.profileDict['Profile']
                # try:
                #     polarionExcelPath = profile['TestCaseExcel']
                #     self.polarionExcelEdit.setText(polarionExcelPath)
                # except KeyError:
                #     self.polarionExcelEdit.setText('')

                # try:
                #     testcasepath = profile['TestCaseExcel']
                #     self.testCaseExcelEdit.setText(testcasepath)
                # except KeyError:
                #     self.testCaseExcelEdit.setText('')

                # try:
                #     callFunctionFolder = profile['CallFunctionFolder']
                #     if callFunctionFolder:
                #         self.callFunctionEdit.setText(callFunctionFolder)
                #         # self.checkFolderExist(callFunctionFolder)
                # except KeyError:
                #     self.callFunctionEdit.setText('')

                # try:
                #     csvPath = profile['CSVReportFolder']
                #     if csvPath:
                #         self.csvReportEdit.setText(csvPath)
                #         self.checkFolderExist(csvPath)
                # except KeyError:
                #     self.csvReportEdit.setText('')

                try:
                    varPoolPath = profile['VariablePoolPath']
                    self.variablePoolEdit.setText(varPoolPath)
                    # self.checkVarPoolExist(varPoolPath)
                except KeyError:
                    self.variablePoolEdit.setText('')

                try:
                    dtcExceptionEnable = profile['DTC']['@enable']
                    self.dtcExCheckBox.setChecked(dtcExceptionEnable)
                    self.showDtcException(dtcExceptionEnable)
                except KeyError:
                    self.dtcExCheckBox.setChecked(False)
                    self.showDtcException(False)
                except TypeError:
                    print('TypeError: eval() arg 1 must be a string or code object')

                try:
                    callFunctionDebug = profile['CallFunctionDebug']['@enable']
                    self.callFunctionDebugCheckbox.setChecked(callFunctionDebug)
                except KeyError:
                    self.callFunctionDebugCheckbox.setChecked(False)

                try:
                    versionCheckbox = profile['Version']['@include']
                    self.versionCheckBox.setChecked(versionCheckbox)
                except KeyError:
                    self.versionCheckBox.setChecked(False)

                # try:
                #     useGuiRunList = profile['UseGuiRunList']
                #     self.useGuiRunListCheckbox.setChecked(useGuiRunList)
                # except KeyError:
                #     self.useGuiRunListCheckbox.setChecked(False)

                try:
                    self.pauseCheckbox.setChecked(profile['PauseEnable'])
                except:
                    self.pauseCheckbox.setChecked(False)

                try:
                    self.pauseStepSpinbox.setValue(int(profile['PauseStep']))
                except:
                    self.pauseStepSpinbox.setValue(1)

                try:
                    self.printStepsInOutputCheckbox.setChecked(profile['PrintStepsInOutput'])
                except:
                    self.printStepsInOutputCheckbox.setChecked(False)

                try:
                    self.maxPrintsInOutputSpinbox.setValue(profile['MaxPrintsInOutput'])
                except:
                    self.maxPrintsInOutputSpinbox.setValue(5)

                try:
                    self.titleCheckBox.setChecked(profile['IncludeTitle'])
                except:
                    self.titleCheckBox.setChecked(False)

                try:
                    self.descriptionCheckBox.setChecked(profile['IncludeDescription'])
                except:
                    self.descriptionCheckBox.setChecked(False)

                self.addSignalModel.clear()
                try:
                    templist = profile['Log']['Signal']
                    signalList = []
                    # if temp is a list we iterate and add, else we append a single element
                    if isinstance(templist, list):
                        signalList.extend(x for x in templist)
                    else:
                        signalList.append(templist)

                    for s in signalList:
                        self.addSignalModel.appendRow(QStandardItem(str(s)))
                except KeyError:
                    debugPrint('Signal list is empty')

                self.updateDtcTableModel()
                self.updateVersionTableModel()
                self.updateBatchTableModel()
        except:
            print(traceback.format_exc())

    def loadSettings(self):
        """Load settings"""
        try:
            self.settings = QSettings('Karma', 'AutomationGUI')
            self.profileFile = self.settings.value('LastProfile', type=str)
            autoRun = self.settings.value('AutoRun', type=bool)
            self.autorunCheckBox.setChecked(autoRun)
            self.autorunSpinBox.setValue(self.settings.value('AutoRunTimer', type=int))
            width = self.settings.value('Width', type=int)
            height = self.settings.value('Height', type=int)
            self.resize(1280 if width > 1900 else width, 720 if height > 1000 else height)

            self.polarionUsername = self.settings.value('PolarionUsername', type=str)

            polarionLeftTableWidth = self.settings.value('PolarionLeftTableWidth', type=int)

            self.polarionTableView.resize(polarionLeftTableWidth, self.polarionTableView.height())

            if os.path.exists(self.profileFile):
                self.loadProfile()
        except:
            print(traceback.format_exc())

    def loadVariablePool(self):
        """Load variable pool from file"""
        try:
            varpoolfile = Path(self.variablePoolEdit.text())
            if os.path.exists(str(varpoolfile)):
                try:
                    with open(str(varpoolfile)) as f:
                        # self.variablePool.extend(row['VariableName'] for row in csv.DictReader(f))
                        self.variablePoolDict = eval(f.read())

                        variablePoolKeys = list(self.variablePoolDict)

                        model = QStringListModel()
                        model.setStringList(variablePoolKeys)
                        completer = QCompleter()
                        completer.setModel(model)
                        completer.setCaseSensitivity(0)
                        self.addSignalEdit.setCompleter(completer)
                        self.statusbar.showMessage('Variable pool loaded...OK!')
                except FileNotFoundError:
                    self.statusbar.showMessage('Variable pool file not found')
        except:
            print(traceback.format_exc())

    def setTitle(self, profilePath):
        """Set window title"""
        try:
            self.setWindowTitle('[' + str(profilePath) + '] - AutomationDesk GUI ' + version)
        except:
            print(traceback.format_exc())

    def exit(self):
        """Exit"""
        try:
            if self.unsavedChanges:
                msgReply = QMessageBox.question(
                    self,
                    'Save Changes',
                    'Would you like to save before exit?',
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )

                if msgReply == QMessageBox.Yes:
                    self.saveProfile()
                    debugPrint('File saved')

            self.saveSettings()
            self.makeRunList()
            self.getExcelData()
            self.updateProfileFromGui()
            # self.saveProfile()
            self.close()
        except:
            print(traceback.format_exc())

    def about(self):
        """About dialog"""
        try:
            about = QMessageBox()
            about.setWindowIcon(QIcon(':/icon/karmalogo24'))
            about.setWindowTitle('About')
            about.setText(version + '\n'
                                    'Developer: Vu Le\n'
                                    'The developer reserves all rights to this software.\n'
                                    'Do not distribute this software without permission')
            about.setInformativeText('Copyright (C) 2018')
            pixmap = QPixmap(':/icon/karmalogolarge')
            # pixmap = pixmap.scaledToWidth(128)
            about.setIconPixmap(pixmap)
            about.exec_()
        except:
            print(traceback.format_exc())

def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')

    # app.setAttribute(Qt.AA_UseHighDpiPixmaps)

    dark_palette = QPalette()
    dark_palette.setColor(QPalette.Window, QColor(53, 53, 53))
    dark_palette.setColor(QPalette.WindowText, Qt.white)
    dark_palette.setColor(QPalette.Base, QColor(53, 53, 53))
    dark_palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
    dark_palette.setColor(QPalette.ToolTipBase, Qt.white)
    dark_palette.setColor(QPalette.ToolTipText, Qt.white)
    dark_palette.setColor(QPalette.Text, Qt.white)
    dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))
    dark_palette.setColor(QPalette.ButtonText, Qt.white)
    dark_palette.setColor(QPalette.BrightText, Qt.red)
    dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
    dark_palette.setColor(QPalette.Highlight, QColor(138, 55, 11))
    dark_palette.setColor(QPalette.HighlightedText, Qt.white)
    app.setPalette(dark_palette)
    app.setStyleSheet("QToolTip { color: #ffffff; background-color: #2a82da; border: 1px solid white; }")

    form = App()
    form.show()
    app.exec_()
    return form.excelData, form.runList, form.profileDict, form.variablePoolDict  # return data to AutomationDesk

if __name__ == '__main__':
    main()